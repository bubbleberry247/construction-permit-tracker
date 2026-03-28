"""
既存CSV/Excel → SQLite 移行スクリプト

使用法:
    python src/migrate.py
"""
import csv
import json
import re
import sqlite3
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

sys.stdout.reconfigure(encoding="utf-8")

# ---------------------------------------------------------------------------
# パス定義
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).parent.parent
SRC_DIR = Path(__file__).parent

# db.pyをインポート
sys.path.insert(0, str(SRC_DIR))
from db import DB_PATH, get_connection, init_db

COMPANY_MASTER = PROJECT_ROOT / "output" / "company_master.csv"
SENT_EMAILS = Path("C:/tmp/sent_emails_raw.txt")
NEW_VENDORS = Path("C:/tmp/new_vendors_from_sent.csv")
INBOUND_LOG = PROJECT_ROOT / "logs" / "inbound_log.csv"
PAGE_CLASS = Path("C:/tmp/page_classification.csv")
OCR_RESULTS = Path("C:/tmp/permit_ocr_results.csv")
REOCR_RESULTS = Path("C:/tmp/permit_reocr_results.json")
FILE_MAPPING = Path("C:/tmp/file_mapping_final.csv")
ROTATION_LOG = Path("C:/tmp/rotation_log.csv")
EXCEL_145 = Path("C:/tmp/継続取引業者リスト_145社.xlsx")

# ---------------------------------------------------------------------------
# ユーザー手動修正値
# ---------------------------------------------------------------------------
MANUAL_CORRECTIONS: list[dict[str, Any]] = [
    {"company": "有限会社キー・カンパニー", "permit_number": "57897", "expiry_date": "2031-02-17",
     "trades": "電気工事業,管工事業", "permit_authority": "愛知県知事", "permit_category": "般"},
    {"company": "株式会社井上商会", "permit_number": "36304", "expiry_date": "2030-11-24",
     "trades": "", "permit_authority": "愛知県知事", "permit_category": "特"},
    {"company": "株式会社オカケン", "permit_number": "85080", "expiry_date": "2028-06-07",
     "trades": "とび・土工工事業", "permit_authority": "愛知県知事", "permit_category": "般"},
    {"company": "株式会社鈴木産業", "permit_number": "28652", "expiry_date": "2030-11-27",
     "trades": "", "permit_authority": "国土交通大臣", "permit_category": "特"},
    {"company": "株式会社メンテック", "permit_number": "69212", "expiry_date": "2026-09-20",
     "trades": "塗装工事業", "permit_authority": "愛知県知事", "permit_category": "般"},
    {"company": "福釜電工株式会社", "permit_number": "45685", "expiry_date": "2030-10-14",
     "trades": "電気工事業", "permit_authority": "愛知県知事", "permit_category": "特"},
    {"company": "有限会社 竹下建築", "permit_number": "14771", "expiry_date": "2027-11-19",
     "trades": "", "permit_authority": "岐阜県知事", "permit_category": "般"},
]


def _read_csv(path: Path, encoding: str = "utf-8") -> list[dict[str, str]]:
    """CSV読み込み（BOM対応）"""
    with open(path, "r", encoding=encoding) as f:
        reader = csv.DictReader(f)
        return list(reader)


def _read_csv_sig(path: Path) -> list[dict[str, str]]:
    return _read_csv(path, encoding="utf-8-sig")


# ===========================================================================
# Step 1: companies
# ===========================================================================

def migrate_companies(conn: sqlite3.Connection) -> dict[str, str]:
    """company_master.csv → companies テーブル。name→company_id マップを返す"""
    print("\n[1/8] companies ...")
    rows = _read_csv_sig(COMPANY_MASTER)
    name_to_id: dict[str, str] = {}

    for r in rows:
        status = r.get("status", "ACTIVE")
        cid = r["company_id"]
        name = r["official_name"]
        conn.execute(
            """INSERT OR IGNORE INTO companies
               (company_id, official_name, name_aliases, corporation_type,
                permit_number, permit_authority, mlit_status,
                last_confirmed_at, status, created_at, updated_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (cid, name, r.get("name_aliases", ""),
             r.get("corporation_type", ""),
             r.get("permit_number", ""), r.get("permit_authority", ""),
             r.get("mlit_status", "NOT_CONFIRMED"),
             r.get("last_confirmed_at", ""), status,
             r.get("created_at", ""), r.get("updated_at", "")),
        )
        name_to_id[name] = cid
        # エイリアスもマップ
        for alias in (r.get("name_aliases") or "").split("|"):
            alias = alias.strip()
            if alias:
                name_to_id[alias] = cid

    # new_vendors_from_sent.csv の新規企業
    if NEW_VENDORS.exists():
        new_rows = _read_csv(NEW_VENDORS)
        next_id = max(int(cid[1:]) for cid in name_to_id.values() if cid.startswith("C")) + 1
        for r in new_rows:
            official = r.get("official_name", "").strip()
            if not official:
                continue
            if official in name_to_id:
                continue
            cid = f"C{next_id:04d}"
            next_id += 1
            conn.execute(
                """INSERT OR IGNORE INTO companies
                   (company_id, official_name, status, created_at, updated_at)
                   VALUES (?,?,?,datetime('now','localtime'),datetime('now','localtime'))""",
                (cid, official, "ACTIVE"),
            )
            name_to_id[official] = cid

    cnt = conn.execute("SELECT count(*) FROM companies").fetchone()[0]
    print(f"  -> {cnt} 社")
    conn.commit()
    return name_to_id


# ===========================================================================
# Step 2: company_emails
# ===========================================================================

def migrate_company_emails(conn: sqlite3.Connection, name_to_id: dict[str, str]) -> None:
    print("\n[2/8] company_emails ...")
    inserted = 0

    # company_master.csv の sender_emails
    rows = _read_csv_sig(COMPANY_MASTER)
    for r in rows:
        cid = r["company_id"]
        for email in (r.get("sender_emails") or "").split("|"):
            email = email.strip()
            if email:
                try:
                    conn.execute(
                        "INSERT OR IGNORE INTO company_emails (company_id, email, email_type) VALUES (?,?,?)",
                        (cid, email, "sender"),
                    )
                    inserted += 1
                except sqlite3.IntegrityError:
                    pass

    # sent_emails_raw.txt → recipient type
    if SENT_EMAILS.exists():
        with open(SENT_EMAILS, "r", encoding="utf-8") as f:
            for line in f:
                email = line.strip()
                if not email or "@" not in email:
                    continue
                # メールアドレスで会社を特定
                found_cid = _find_company_by_email(conn, email, name_to_id)
                if found_cid:
                    try:
                        conn.execute(
                            "INSERT OR IGNORE INTO company_emails (company_id, email, email_type) VALUES (?,?,?)",
                            (found_cid, email, "recipient"),
                        )
                        inserted += 1
                    except sqlite3.IntegrityError:
                        pass

    print(f"  -> {inserted} 件")
    conn.commit()


def _find_company_by_email(conn: sqlite3.Connection, email: str, name_to_id: dict[str, str]) -> str | None:
    """メールアドレスから企業IDを逆引き"""
    # company_emails テーブルで検索
    row = conn.execute(
        "SELECT company_id FROM company_emails WHERE email = ?", (email,)
    ).fetchone()
    if row:
        return row[0]
    return None


# ===========================================================================
# Step 3: inbound_messages + files
# ===========================================================================

def migrate_inbound(conn: sqlite3.Connection, name_to_id: dict[str, str]) -> dict[str, int]:
    """inbound_log.csv → inbound_messages + files"""
    print("\n[3/8] inbound_messages + files ...")
    rows = _read_csv_sig(INBOUND_LOG)
    file_hash_to_id: dict[str, int] = {}
    msg_count = 0
    file_count = 0

    # ファイルマッピング（company_name → file info）
    fmap: dict[str, list[dict]] = {}
    if FILE_MAPPING.exists():
        for r in _read_csv_sig(FILE_MAPPING):
            cn = r.get("company_name", "").strip()
            if cn:
                fmap.setdefault(cn, []).append(r)

    seen_messages: set[str] = set()
    for r in rows:
        mid = r.get("message_id", "").strip()
        if not mid or mid in seen_messages:
            continue
        seen_messages.add(mid)

        sender = r.get("sender_email", "").strip()
        original_sender = r.get("original_sender_email", "").strip()
        received = r.get("received_at", "").strip()

        # 企業マッチ: original_sender > sender でcompany_emailsからルックアップ
        company_id = None
        for em in [original_sender, sender]:
            if em:
                row2 = conn.execute(
                    "SELECT company_id FROM company_emails WHERE email = ?", (em,)
                ).fetchone()
                if row2:
                    company_id = row2[0]
                    break

        conn.execute(
            """INSERT OR IGNORE INTO inbound_messages
               (message_id, company_id, sender_email, original_sender, received_at)
               VALUES (?,?,?,?,?)""",
            (mid, company_id, sender, original_sender, received),
        )
        msg_count += 1

        # files
        fname = r.get("file_name", "").strip()
        fhash = r.get("file_hash", "").strip()
        fsize = r.get("file_size_bytes", "").strip()
        spath = r.get("saved_path", "").strip()

        if fname:
            try:
                try:
                    fsize_int = int(fsize) if fsize and fsize.isdigit() else None
                except (ValueError, TypeError):
                    fsize_int = None
                cur = conn.execute(
                    """INSERT INTO files
                       (message_id, company_id, file_name, file_hash, file_size_bytes, saved_path)
                       VALUES (?,?,?,?,?,?)""",
                    (mid, company_id, fname, fhash,
                     fsize_int, spath),
                )
                file_id = cur.lastrowid
                if fhash:
                    file_hash_to_id[fhash] = file_id
                file_count += 1
            except sqlite3.IntegrityError as e:
                print(f"  [WARN] file重複: {fname} - {e}")

    # file_mapping_final.csv の追加情報をfilesに更新
    if FILE_MAPPING.exists():
        for r in _read_csv_sig(FILE_MAPPING):
            fhash = r.get("file_hash", "").strip()
            if fhash and fhash in file_hash_to_id:
                conn.execute(
                    "UPDATE files SET new_filename=?, new_path=? WHERE file_id=?",
                    (r.get("new_filename", ""), r.get("new_path", ""),
                     file_hash_to_id[fhash]),
                )
            elif fhash:
                # hashがinbound_logに無いファイル → file_mappingのみのエントリ
                cn = r.get("company_name", "").strip()
                cid = name_to_id.get(cn)
                try:
                    fsize_str = r.get("file_size_kb", "").strip()
                    fsize_bytes = int(float(fsize_str) * 1024) if fsize_str else None
                except ValueError:
                    fsize_bytes = None
                cur = conn.execute(
                    """INSERT INTO files
                       (company_id, file_name, file_hash, file_size_bytes,
                        saved_path, new_filename, new_path)
                       VALUES (?,?,?,?,?,?,?)""",
                    (cid, r.get("new_filename", ""), fhash, fsize_bytes,
                     r.get("original_path", ""),
                     r.get("new_filename", ""), r.get("new_path", "")),
                )
                file_hash_to_id[fhash] = cur.lastrowid
                file_count += 1

    print(f"  -> messages: {msg_count}, files: {file_count}")
    conn.commit()
    return file_hash_to_id


# ===========================================================================
# Step 4: pages
# ===========================================================================

def migrate_pages(conn: sqlite3.Connection, file_hash_to_id: dict[str, int],
                  name_to_id: dict[str, str]) -> dict[tuple[str, int], int]:
    """page_classification.csv + rotation_log.csv → pages"""
    print("\n[4/8] pages ...")
    page_key_to_id: dict[tuple[str, int], int] = {}  # (filename, page_no) -> page_id

    # rotation情報をロード
    rotations: dict[tuple[str, int], int] = {}
    if ROTATION_LOG.exists():
        for r in _read_csv(ROTATION_LOG):
            fn = r.get("filename", "").strip()
            try:
                pno = int(r.get("page_no", "0"))
                rot = int(r.get("final_rotation", "0"))
                rotations[(fn, pno)] = rot
            except ValueError:
                pass

    if PAGE_CLASS.exists():
        rows = _read_csv_sig(PAGE_CLASS)
        for r in rows:
            company = r.get("company", "").strip()
            fn = r.get("filename", "").strip()
            fhash = r.get("file_hash", "").strip()
            try:
                pno = int(r.get("page_no", "0"))
            except ValueError:
                pno = 0
            doc_type_id_str = r.get("doc_type_id", "").strip()
            doc_type_name = r.get("doc_type_name", "").strip()
            conf_str = r.get("confidence", "").strip()

            file_id = file_hash_to_id.get(fhash)
            cid = name_to_id.get(company)

            try:
                doc_tid = int(doc_type_id_str) if doc_type_id_str else None
            except ValueError:
                doc_tid = None

            try:
                conf = float(conf_str) if conf_str else None
            except ValueError:
                conf = None

            rotation = rotations.get((fn, pno), 0)

            cur = conn.execute(
                """INSERT INTO pages
                   (file_id, company_id, file_name, file_hash, page_no,
                    doc_type_id, doc_type_name, confidence, rotation)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (file_id, cid, fn, fhash, pno,
                 doc_tid, doc_type_name, conf, rotation),
            )
            page_key_to_id[(fn, pno)] = cur.lastrowid

    cnt = conn.execute("SELECT count(*) FROM pages").fetchone()[0]
    print(f"  -> {cnt} ページ")
    conn.commit()
    return page_key_to_id


# ===========================================================================
# Step 5: ocr_runs + ocr_fields
# ===========================================================================

def migrate_ocr(conn: sqlite3.Connection, page_key_to_id: dict[tuple[str, int], int],
                name_to_id: dict[str, str]) -> dict[str, int]:
    """permit_ocr_results.csv + reocr → ocr_runs + ocr_fields"""
    print("\n[5/8] ocr_runs + ocr_fields ...")
    company_run_map: dict[str, int] = {}  # company_name -> run_id

    # 初回OCR結果
    if OCR_RESULTS.exists():
        rows = _read_csv_sig(OCR_RESULTS)
        for r in rows:
            company = r.get("company", "").strip()
            fn = r.get("filename", "").strip()
            try:
                pno = int(r.get("page_no", "0"))
            except ValueError:
                pno = 0

            page_id = page_key_to_id.get((fn, pno))

            raw_resp = r.get("raw_response", "")
            cur = conn.execute(
                """INSERT INTO ocr_runs (page_id, run_type, model_name, status, raw_response)
                   VALUES (?, 'initial', 'gpt-4o', 'completed', ?)""",
                (page_id, raw_resp),
            )
            run_id = cur.lastrowid
            company_run_map[company] = run_id

            # フィールドを個別に保存
            field_names = [
                "permit_number", "permit_authority", "permit_category",
                "permit_year", "expiry_date", "issue_date", "trade_categories",
            ]
            for fn_key in field_names:
                val = r.get(fn_key, "").strip()
                if val:
                    conn.execute(
                        """INSERT INTO ocr_fields (run_id, field_name, raw_value, normalized)
                           VALUES (?,?,?,?)""",
                        (run_id, fn_key, val, val),
                    )

    # 再OCR結果
    if REOCR_RESULTS.exists():
        with open(REOCR_RESULTS, "r", encoding="utf-8") as f:
            reocr = json.load(f)
        for item in reocr:
            company = item.get("company", "").strip()
            page_id = None  # reocr結果にはpage情報がないため

            cur = conn.execute(
                """INSERT INTO ocr_runs (page_id, run_type, model_name, status, raw_response)
                   VALUES (?, 'reocr', 'gpt-4o', 'completed', ?)""",
                (page_id, json.dumps(item, ensure_ascii=False)),
            )
            run_id = cur.lastrowid
            company_run_map[company] = run_id  # 上書き（最新優先）

            field_names = [
                "permit_number", "permit_authority", "permit_category",
                "permit_year", "expiry_date", "issue_date", "trade_categories",
            ]
            for fn_key in field_names:
                val = str(item.get(fn_key, "")).strip()
                if val:
                    conn.execute(
                        """INSERT INTO ocr_fields (run_id, field_name, raw_value, normalized)
                           VALUES (?,?,?,?)""",
                        (run_id, fn_key, val, val),
                    )

    run_cnt = conn.execute("SELECT count(*) FROM ocr_runs").fetchone()[0]
    field_cnt = conn.execute("SELECT count(*) FROM ocr_fields").fetchone()[0]
    print(f"  -> runs: {run_cnt}, fields: {field_cnt}")
    conn.commit()
    return company_run_map


# ===========================================================================
# Step 6: field_reviews (ユーザー手動修正)
# ===========================================================================

def migrate_field_reviews(conn: sqlite3.Connection, name_to_id: dict[str, str]) -> None:
    print("\n[6/8] field_reviews (手動修正) ...")
    inserted = 0

    for mc in MANUAL_CORRECTIONS:
        company = mc["company"]
        cid = name_to_id.get(company)
        if not cid:
            # 部分一致で探す
            for name, id_ in name_to_id.items():
                if company in name or name in company:
                    cid = id_
                    break
        if not cid:
            print(f"  [WARN] 企業不明: {company}")
            continue

        fields = {
            "permit_number": mc.get("permit_number"),
            "expiry_date": mc.get("expiry_date"),
            "permit_authority": mc.get("permit_authority"),
            "permit_category": mc.get("permit_category"),
        }
        if mc.get("trades"):
            fields["trade_categories"] = mc["trades"]

        for fn, val in fields.items():
            if val:
                conn.execute(
                    """INSERT INTO field_reviews
                       (company_id, field_name, confirmed_value, confirmed_by)
                       VALUES (?,?,?,?)""",
                    (cid, fn, val, "user_manual"),
                )
                inserted += 1

    print(f"  -> {inserted} 件")
    conn.commit()


# ===========================================================================
# Step 7: permits + permit_trades
# ===========================================================================

def migrate_permits(conn: sqlite3.Connection, name_to_id: dict[str, str]) -> None:
    """OCR結果 + 手動修正 → permits + permit_trades"""
    print("\n[7/8] permits + permit_trades ...")

    # 1) OCR結果ベースの許可証
    if OCR_RESULTS.exists():
        rows = _read_csv_sig(OCR_RESULTS)
        for r in rows:
            company = r.get("company", "").strip()
            cid = _resolve_company(company, name_to_id)
            if not cid:
                print(f"  [WARN] 企業不明: {company}")
                continue

            pn = r.get("permit_number", "").strip()
            if not pn or pn == "UNCERTAIN":
                continue

            cur = conn.execute(
                """INSERT INTO permits
                   (company_id, permit_number, permit_authority, permit_category,
                    permit_year, issue_date, expiry_date, source)
                   VALUES (?,?,?,?,?,?,?,?)""",
                (cid, pn,
                 r.get("permit_authority", ""),
                 r.get("permit_category", ""),
                 r.get("permit_year", ""),
                 r.get("issue_date", ""),
                 r.get("expiry_date", ""),
                 "ocr"),
            )
            permit_id = cur.lastrowid
            _insert_trades(conn, permit_id, r.get("trade_categories", ""))

    # 2) 再OCR結果（初回と重複する場合は上書き: current_flag=0にして新規追加）
    if REOCR_RESULTS.exists():
        with open(REOCR_RESULTS, "r", encoding="utf-8") as f:
            reocr = json.load(f)
        for item in reocr:
            company = item.get("company", "").strip()
            cid = _resolve_company(company, name_to_id)
            if not cid:
                print(f"  [WARN] reocr企業不明: {company}")
                continue

            pn = str(item.get("permit_number", "")).strip()
            if not pn or pn == "UNCERTAIN":
                continue

            # 同じ会社の旧OCR結果をcurrent_flag=0に
            conn.execute(
                "UPDATE permits SET current_flag=0 WHERE company_id=? AND source='ocr'",
                (cid,),
            )

            expiry = str(item.get("expiry_date", "")).strip()
            cur = conn.execute(
                """INSERT INTO permits
                   (company_id, permit_number, permit_authority, permit_category,
                    permit_year, issue_date, expiry_date, source)
                   VALUES (?,?,?,?,?,?,?,?)""",
                (cid, pn,
                 str(item.get("permit_authority", "")),
                 str(item.get("permit_category", "")),
                 str(item.get("permit_year", "")),
                 str(item.get("issue_date", "")),
                 expiry,
                 "reocr"),
            )
            permit_id = cur.lastrowid
            _insert_trades(conn, permit_id, str(item.get("trade_categories", "")))

    # 3) ユーザー手動修正（最優先）
    for mc in MANUAL_CORRECTIONS:
        company = mc["company"]
        cid = _resolve_company(company, name_to_id)
        if not cid:
            continue

        pn = mc.get("permit_number", "")
        if not pn:
            continue

        # 同じ会社の既存をcurrent_flag=0に
        conn.execute(
            "UPDATE permits SET current_flag=0 WHERE company_id=?",
            (cid,),
        )

        cur = conn.execute(
            """INSERT INTO permits
               (company_id, permit_number, permit_authority, permit_category,
                issue_date, expiry_date, source)
               VALUES (?,?,?,?,?,?,?)""",
            (cid, pn,
             mc.get("permit_authority", ""),
             mc.get("permit_category", ""),
             "",
             mc.get("expiry_date", ""),
             "manual"),
        )
        permit_id = cur.lastrowid
        _insert_trades(conn, permit_id, mc.get("trades", ""))

    permit_cnt = conn.execute("SELECT count(*) FROM permits").fetchone()[0]
    trade_cnt = conn.execute("SELECT count(*) FROM permit_trades").fetchone()[0]
    current_cnt = conn.execute("SELECT count(*) FROM permits WHERE current_flag=1").fetchone()[0]
    print(f"  -> permits: {permit_cnt} (current: {current_cnt}), trades: {trade_cnt}")
    conn.commit()


def _resolve_company(name: str, name_to_id: dict[str, str]) -> str | None:
    """企業名からcompany_idを解決（完全一致→部分一致）"""
    if name in name_to_id:
        return name_to_id[name]
    # 部分一致
    for key, cid in name_to_id.items():
        if name in key or key in name:
            return cid
    return None


def _insert_trades(conn: sqlite3.Connection, permit_id: int, trades_str: str) -> None:
    """カンマ区切りの業種文字列をpermit_tradesに挿入"""
    if not trades_str:
        return
    for trade in trades_str.split(","):
        trade = trade.strip()
        if trade and trade != "UNCERTAIN":
            conn.execute(
                "INSERT INTO permit_trades (permit_id, trade_name) VALUES (?,?)",
                (permit_id, trade),
            )


# ===========================================================================
# Step 8: 差分レポート
# ===========================================================================

def diff_report(conn: sqlite3.Connection) -> None:
    """v_current_permits vs Excel比較"""
    print("\n[8/8] 差分レポート: v_current_permits vs Excel ...")

    try:
        import openpyxl
    except ImportError:
        print("  [ERROR] openpyxl未インストール: pip install openpyxl")
        return

    if not EXCEL_145.exists():
        print(f"  [ERROR] Excelファイル不在: {EXCEL_145}")
        return

    # DB側
    db_permits: dict[str, dict] = {}
    for row in conn.execute("SELECT * FROM v_current_permits").fetchall():
        name = row["official_name"]
        db_permits[name] = {
            "permit_number": row["permit_number"],
            "permit_authority": row["permit_authority"],
            "permit_category": row["permit_category"],
            "expiry_date": row["expiry_date"],
            "trade_names": row["trade_names"],
            "source": row["source"],
        }

    # Excel側
    wb = openpyxl.load_workbook(str(EXCEL_145), read_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    excel_permits: dict[str, dict] = {}
    seen_names: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0] or "").strip()
        if not name or "@" in name:  # メールアドレス行をスキップ
            continue
        if name in seen_names:
            continue
        seen_names.add(name)

        permit_num = str(row[5] or "").strip()
        permit_auth = str(row[8] or "").strip()
        trade = str(row[9] or "").strip()
        expiry = str(row[10] or "").strip()
        status = str(row[14] or "").strip()

        # Y-AB列
        pn2 = str(row[24] or "").strip() if len(row) > 24 else ""
        pa2 = str(row[25] or "").strip() if len(row) > 25 else ""
        cat2 = str(row[26] or "").strip() if len(row) > 26 else ""
        exp2 = str(row[27] or "").strip() if len(row) > 27 else ""

        excel_permits[name] = {
            "permit_number": pn2 or permit_num,
            "permit_authority": pa2 or permit_auth,
            "permit_category": cat2,
            "expiry_date": exp2 or expiry,
            "trade": trade,
            "status": status,
        }
    wb.close()

    # 差分出力
    print(f"\n{'='*80}")
    print(f"  DB側: {len(db_permits)} 社  |  Excel側: {len(excel_permits)} 社")
    print(f"{'='*80}")

    # DB にあって Excel にない
    db_only = set(db_permits.keys()) - set(excel_permits.keys())
    if db_only:
        print(f"\n--- DBのみ ({len(db_only)} 社) ---")
        for name in sorted(db_only):
            d = db_permits[name]
            print(f"  {name}: 許可#{d['permit_number']} ({d['source']})")

    # Excel にあって DB にない（有効許可あり）
    excel_only_with_permit = []
    for name in sorted(set(excel_permits.keys()) - set(db_permits.keys())):
        e = excel_permits[name]
        if e["permit_number"]:
            excel_only_with_permit.append((name, e))
    if excel_only_with_permit:
        print(f"\n--- Excelのみ（許可番号あり） ({len(excel_only_with_permit)} 社) ---")
        for name, e in excel_only_with_permit:
            print(f"  {name}: 許可#{e['permit_number']} 有効期限={e['expiry_date']}")

    # 両方にある → 許可番号の不一致
    both = set(db_permits.keys()) & set(excel_permits.keys())
    diffs = []
    for name in sorted(both):
        d = db_permits[name]
        e = excel_permits[name]

        d_pn = (d["permit_number"] or "").lstrip("0")
        e_pn = (e["permit_number"] or "").lstrip("0")

        if d_pn and e_pn and d_pn != e_pn:
            diffs.append((name, d, e))

    if diffs:
        print(f"\n--- 許可番号不一致 ({len(diffs)} 社) ---")
        for name, d, e in diffs:
            print(f"  {name}:")
            print(f"    DB:    #{d['permit_number']} expire={d['expiry_date']} src={d['source']}")
            print(f"    Excel: #{e['permit_number']} expire={e['expiry_date']}")

    # 有効期限の差異
    expiry_diffs = []
    for name in sorted(both):
        d = db_permits[name]
        e = excel_permits[name]

        d_exp = (d["expiry_date"] or "").strip()
        e_exp = (e["expiry_date"] or "").strip()

        # 日付正規化: datetimeオブジェクト文字列をYYYY-MM-DDに
        if "00:00:00" in e_exp:
            e_exp = e_exp.split(" ")[0]

        if d_exp and e_exp and d_exp != "UNCERTAIN" and d_exp != e_exp:
            expiry_diffs.append((name, d_exp, e_exp))

    if expiry_diffs:
        print(f"\n--- 有効期限不一致 ({len(expiry_diffs)} 社) ---")
        for name, d_exp, e_exp in expiry_diffs:
            print(f"  {name}: DB={d_exp} vs Excel={e_exp}")

    # サマリ
    print(f"\n{'='*80}")
    print(f"  DBのみ: {len(db_only)} 社")
    print(f"  Excelのみ(許可あり): {len(excel_only_with_permit)} 社")
    print(f"  許可番号不一致: {len(diffs)} 社")
    print(f"  有効期限不一致: {len(expiry_diffs)} 社")
    print(f"{'='*80}")


# ===========================================================================
# メイン
# ===========================================================================

def main() -> None:
    print("=" * 60)
    print(" 建設業許可証管理DB 移行開始")
    print("=" * 60)

    # 既存DB削除して再作成
    if DB_PATH.exists():
        DB_PATH.unlink()
        print(f"[既存DB削除] {DB_PATH}")

    init_db()
    conn = get_connection()

    try:
        name_to_id = migrate_companies(conn)
        migrate_company_emails(conn, name_to_id)
        file_hash_to_id = migrate_inbound(conn, name_to_id)
        page_key_to_id = migrate_pages(conn, file_hash_to_id, name_to_id)
        migrate_ocr(conn, page_key_to_id, name_to_id)
        migrate_field_reviews(conn, name_to_id)
        migrate_permits(conn, name_to_id)
        diff_report(conn)
    finally:
        conn.close()

    print(f"\n[完了] DB: {DB_PATH}")
    print(f"  サイズ: {DB_PATH.stat().st_size / 1024:.1f} KB")


if __name__ == "__main__":
    main()
