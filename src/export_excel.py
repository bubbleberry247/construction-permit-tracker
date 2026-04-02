"""
export_excel.py — 建設業許可証管理台帳 Excel出力（4シート構成）

Sheet 1: 書類受領状況  — 9書類の受領○/—チェック一覧
Sheet 2: 許可証情報    — 全許可証の詳細情報
Sheet 3: 期限警告      — 365日以内に満了する許可証
Sheet 4: 許可業種一覧  — 29業種マスタ

Usage:
    python -X utf8 src/export_excel.py
    python -X utf8 src/export_excel.py --output custom_name.xlsx
"""
from __future__ import annotations

import argparse
import sqlite3
import sys
from datetime import date, timedelta
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

# ---------------------------------------------------------------------------
# パス定義
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).parent.parent
DB_PATH = PROJECT_ROOT / "data" / "permit_tracker.db"
OUTPUT_DIR = PROJECT_ROOT / "output"

TODAY = date.today()

# ---------------------------------------------------------------------------
# 29業種 短縮名マッピング
# ---------------------------------------------------------------------------
TRADE_SHORT: dict[str, str] = {
    "土木工事業": "土木",
    "建築工事業": "建築",
    "大工工事業": "大工",
    "左官工事業": "左官",
    "とび・土工工事業": "とび・土工",
    "石工事業": "石",
    "屋根工事業": "屋根",
    "電気工事業": "電気",
    "管工事業": "管",
    "タイル・れんが・ブロック工事業": "タイル",
    "鋼構造物工事業": "鋼構造物",
    "鉄筋工事業": "鉄筋",
    "舗装工事業": "舗装",
    "しゅんせつ工事業": "しゅんせつ",
    "板金工事業": "板金",
    "ガラス工事業": "ガラス",
    "塗装工事業": "塗装",
    "防水工事業": "防水",
    "内装仕上工事業": "内装仕上",
    "機械器具設置工事業": "機械器具",
    "熱絶縁工事業": "熱絶縁",
    "電気通信工事業": "電気通信",
    "造園工事業": "造園",
    "さく井工事業": "さく井",
    "建具工事業": "建具",
    "水道施設工事業": "水道施設",
    "消防施設工事業": "消防施設",
    "清掃施設工事業": "清掃施設",
    "解体工事業": "解体",
}

# 29業種マスタ (No, コード, 正式名称, 略称)
TRADE_29_MASTER: list[tuple[int, str, str, str]] = [
    (1, "010", "土木工事業", "土木"),
    (2, "020", "建築工事業", "建築"),
    (3, "030", "大工工事業", "大工"),
    (4, "040", "左官工事業", "左官"),
    (5, "050", "とび・土工工事業", "とび・土工"),
    (6, "060", "石工事業", "石"),
    (7, "070", "屋根工事業", "屋根"),
    (8, "080", "電気工事業", "電気"),
    (9, "090", "管工事業", "管"),
    (10, "100", "タイル・れんが・ブロック工事業", "タイル"),
    (11, "110", "鋼構造物工事業", "鋼構造物"),
    (12, "120", "鉄筋工事業", "鉄筋"),
    (13, "130", "舗装工事業", "舗装"),
    (14, "140", "しゅんせつ工事業", "しゅんせつ"),
    (15, "150", "板金工事業", "板金"),
    (16, "160", "ガラス工事業", "ガラス"),
    (17, "170", "塗装工事業", "塗装"),
    (18, "180", "防水工事業", "防水"),
    (19, "190", "内装仕上工事業", "内装仕上"),
    (20, "200", "機械器具設置工事業", "機械器具"),
    (21, "210", "熱絶縁工事業", "熱絶縁"),
    (22, "220", "電気通信工事業", "電気通信"),
    (23, "230", "造園工事業", "造園"),
    (24, "240", "さく井工事業", "さく井"),
    (25, "250", "建具工事業", "建具"),
    (26, "260", "水道施設工事業", "水道施設"),
    (27, "270", "消防施設工事業", "消防施設"),
    (28, "280", "清掃施設工事業", "清掃施設"),
    (29, "290", "解体工事業", "解体"),
]

# 書類種別（Sheet 1 で使用）
DOC_TYPES: list[str] = [
    "取引申請書",
    "建設業許可証",
    "決算書",
    "会社案内",
    "工事経歴書",
    "取引先一覧表",
    "労働安全衛生誓約書",
    "資格略字一覧",
    "労働者名簿",
]

# doc_type_name の表記ゆれ対応 (DB値 → 正規名)
DOC_TYPE_ALIASES: dict[str, str] = {
    "労安誓約書": "労働安全衛生誓約書",
    "資格一覧": "資格略字一覧",
}

# ---------------------------------------------------------------------------
# スタイル定数
# ---------------------------------------------------------------------------
DARK_BLUE = "2B5797"
THIN_SIDE = Side(style="thin", color="000000")
THIN_BORDER = Border(
    left=THIN_SIDE, right=THIN_SIDE,
    top=THIN_SIDE, bottom=THIN_SIDE,
)

HEADER_FILL = PatternFill("solid", fgColor=DARK_BLUE)
HEADER_FONT = Font(name="Yu Gothic UI", bold=True, size=10, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FONT = Font(name="Yu Gothic UI", bold=True, size=14)
SUBTITLE_FONT = Font(name="Yu Gothic UI", bold=False, size=10)
DATA_FONT = Font(name="Yu Gothic UI", size=9)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ステータス色
FILL_GREEN = PatternFill("solid", fgColor="D4EDDA")
FILL_YELLOW = PatternFill("solid", fgColor="FFF3CD")
FILL_RED = PatternFill("solid", fgColor="FFE0E0")
FILL_GRAY = PatternFill("solid", fgColor="F0F0F0")

# ○/— マーク用フォント
FONT_CIRCLE = Font(name="Yu Gothic UI", size=9, color="28A745")  # green
FONT_DASH = Font(name="Yu Gothic UI", size=9, color="DC3545")    # red

# ステータス表示用フォント
FONT_STATUS_GREEN = Font(name="Yu Gothic UI", size=9, bold=True, color="28A745")
FONT_STATUS_YELLOW = Font(name="Yu Gothic UI", size=9, bold=True, color="856404")
FONT_STATUS_GRAY = Font(name="Yu Gothic UI", size=9, color="6C757D")


# ---------------------------------------------------------------------------
# DB接続
# ---------------------------------------------------------------------------
def get_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA foreign_keys = ON")
    conn.row_factory = sqlite3.Row
    return conn


# ---------------------------------------------------------------------------
# DB読み取り
# ---------------------------------------------------------------------------
def load_companies(conn: sqlite3.Connection) -> list[dict[str, str]]:
    """MERGED以外の会社一覧（company_id昇順）"""
    rows = conn.execute(
        "SELECT company_id, official_name FROM companies "
        "WHERE status != 'MERGED' OR status IS NULL "
        "ORDER BY company_id"
    ).fetchall()
    return [dict(r) for r in rows]


def load_company_docs(conn: sqlite3.Connection) -> dict[str, set[str]]:
    """company_id → {doc_type_name, ...}（pages テーブルから）"""
    rows = conn.execute(
        "SELECT DISTINCT company_id, doc_type_name FROM pages "
        "WHERE company_id IS NOT NULL "
        "AND doc_type_name IS NOT NULL AND doc_type_name != ''"
    ).fetchall()
    result: dict[str, set[str]] = {}
    for r in rows:
        doc_name = r["doc_type_name"]
        # 表記ゆれを正規化
        doc_name = DOC_TYPE_ALIASES.get(doc_name, doc_name)
        result.setdefault(r["company_id"], set()).add(doc_name)
    return result


def load_permits(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    """current_flag=1 の許可証一覧"""
    rows = conn.execute(
        """SELECT p.permit_id, p.company_id, p.permit_number,
                  p.permit_authority, p.permit_category, p.permit_year,
                  p.issue_date, p.expiry_date
           FROM permits p
           WHERE p.current_flag = 1
           ORDER BY p.company_id, p.permit_id"""
    ).fetchall()
    return [dict(r) for r in rows]


def load_permit_trades(conn: sqlite3.Connection) -> dict[int, list[str]]:
    """permit_id → [trade_name, ...]"""
    rows = conn.execute(
        "SELECT permit_id, trade_name FROM permit_trades ORDER BY trade_id"
    ).fetchall()
    result: dict[int, list[str]] = {}
    for r in rows:
        result.setdefault(r["permit_id"], []).append(r["trade_name"])
    return result


def load_company_names(conn: sqlite3.Connection) -> dict[str, str]:
    """company_id → official_name"""
    rows = conn.execute("SELECT company_id, official_name FROM companies").fetchall()
    return {r["company_id"]: r["official_name"] for r in rows}


# ---------------------------------------------------------------------------
# ユーティリティ
# ---------------------------------------------------------------------------
def normalize_category(cat: str | None) -> str:
    """許可区分を短縮形に変換: 一般→般, 特定→特"""
    if not cat:
        return ""
    if cat == "一般":
        return "般"
    if cat == "特定":
        return "特"
    return cat  # 既に 般/特 の場合はそのまま


def get_trade_short_names(trade_names: list[str]) -> str:
    """trade_name リストを短縮名のカンマ区切りに変換"""
    shorts: list[str] = []
    for tn in trade_names:
        short = TRADE_SHORT.get(tn)
        if short:
            shorts.append(short)
        else:
            # 部分一致フォールバック（石工工事業 → 石 等）
            matched = False
            for full_name, s in TRADE_SHORT.items():
                key = full_name.replace("工事業", "")
                if key in tn:
                    shorts.append(s)
                    matched = True
                    break
            if not matched:
                shorts.append(tn)
    return "、".join(shorts)


def apply_header_row(
    ws: Any,
    row: int,
    headers: list[str],
    col_widths: list[int | float] | None = None,
) -> None:
    """ヘッダ行を書き込んでスタイル適用"""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    if col_widths:
        for col_idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w


def write_data_cell(
    ws: Any,
    row: int,
    col: int,
    value: Any,
    font: Font | None = None,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
) -> Any:
    """1セル書き込み＋スタイル適用"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = alignment or ALIGN_CENTER
    if fill:
        cell.fill = fill
    return cell


# ---------------------------------------------------------------------------
# Sheet 1: 書類受領状況
# ---------------------------------------------------------------------------
def build_sheet_documents(wb: openpyxl.Workbook, companies: list[dict], company_docs: dict[str, set[str]]) -> None:
    ws = wb.active
    ws.title = "書類受領状況"

    # --- 集計 ---
    total = len(companies)
    received_count = 0   # 1書類以上受信
    complete_count = 0   # 全9書類揃い
    missing_count = 0    # 受信ありだが不足あり
    no_receipt_count = 0 # 未受信

    for c in companies:
        cid = c["company_id"]
        docs = company_docs.get(cid, set())
        matched_docs = sum(1 for dt in DOC_TYPES if dt in docs)
        if matched_docs == 0:
            no_receipt_count += 1
        elif matched_docs >= len(DOC_TYPES):
            complete_count += 1
            received_count += 1
        else:
            missing_count += 1
            received_count += 1

    # --- Row 1: Title ---
    title_cell = ws.cell(row=1, column=1, value="建設業許可証等 提出書類管理台帳")
    title_cell.font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)

    # --- Row 2: Subtitle ---
    subtitle = (
        f"作成日: {TODAY.strftime('%Y-%m-%d')} / "
        f"全{total}社（受信{received_count}社・全揃{complete_count}社"
        f"・不足{missing_count}社・未受信{no_receipt_count}社）"
    )
    subtitle_cell = ws.cell(row=2, column=1, value=subtitle)
    subtitle_cell.font = SUBTITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)

    # --- Row 3: empty (spacer) ---

    # --- Row 4: Headers ---
    headers = [
        "No.", "会社ID", "会社名", "ステータス",
        "取引申請書", "建設業許可証", "決算書", "会社案内",
        "工事経歴書", "取引先一覧表", "労安誓約書", "資格一覧",
        "労働者名簿", "不足書類",
    ]
    col_widths = [5, 8, 30, 10, 10, 12, 8, 8, 10, 12, 10, 8, 10, 40]
    apply_header_row(ws, 4, headers, col_widths)

    # --- Data rows (row 5~) ---
    row_idx = 5
    for seq, c in enumerate(companies, 1):
        cid = c["company_id"]
        name = c["official_name"]
        docs = company_docs.get(cid, set())

        # 各書類の○/—判定
        doc_marks: list[str] = []
        missing_list: list[str] = []
        for dt in DOC_TYPES:
            if dt in docs:
                doc_marks.append("○")
            else:
                doc_marks.append("—")
                missing_list.append(dt)

        has_count = len(DOC_TYPES) - len(missing_list)

        # ステータス判定
        if has_count == 0:
            status = "未受信"
            status_fill = FILL_GRAY
            status_font = FONT_STATUS_GRAY
        elif has_count >= len(DOC_TYPES):
            status = "全揃"
            status_fill = FILL_GREEN
            status_font = FONT_STATUS_GREEN
        else:
            status = "不足"
            status_fill = FILL_YELLOW
            status_font = FONT_STATUS_YELLOW

        missing_text = "、".join(missing_list) if missing_list and has_count > 0 else ""

        # 行書き込み
        write_data_cell(ws, row_idx, 1, seq)
        write_data_cell(ws, row_idx, 2, cid)
        write_data_cell(ws, row_idx, 3, name, alignment=ALIGN_LEFT)
        write_data_cell(ws, row_idx, 4, status, font=status_font, fill=status_fill)

        # 書類列 (col 5-13)
        for i, mark in enumerate(doc_marks):
            if mark == "○":
                write_data_cell(ws, row_idx, 5 + i, mark, font=FONT_CIRCLE)
            else:
                write_data_cell(ws, row_idx, 5 + i, mark, font=FONT_DASH)

        # 不足書類
        write_data_cell(ws, row_idx, 14, missing_text, alignment=ALIGN_LEFT)

        row_idx += 1

    # --- Freeze & Filter ---
    ws.freeze_panes = "A5"
    last_row = row_idx - 1
    if last_row >= 5:
        ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{last_row}"

    return


# ---------------------------------------------------------------------------
# Sheet 2: 許可証情報
# ---------------------------------------------------------------------------
def build_sheet_permits(
    wb: openpyxl.Workbook,
    permits: list[dict[str, Any]],
    company_names: dict[str, str],
    permit_trades: dict[int, list[str]],
) -> None:
    ws = wb.create_sheet("許可証情報")

    # --- Row 1: Title ---
    title_cell = ws.cell(row=1, column=1, value="建設業許可証 一覧")
    title_cell.font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    # --- Row 2: empty (spacer) ---

    # --- Row 3: Headers ---
    headers = [
        "会社ID", "会社名", "許可番号", "許可区分",
        "許可者", "交付日", "有効期限", "許可業種",
    ]
    col_widths = [8, 30, 12, 8, 14, 12, 12, 50]
    apply_header_row(ws, 3, headers, col_widths)

    # --- Data rows (row 4~) ---
    row_idx = 4
    for p in permits:
        cid = p["company_id"]
        name = company_names.get(cid, cid)
        trades = permit_trades.get(p["permit_id"], [])
        trade_text = get_trade_short_names(trades)
        category = normalize_category(p.get("permit_category"))

        write_data_cell(ws, row_idx, 1, cid)
        write_data_cell(ws, row_idx, 2, name, alignment=ALIGN_LEFT)
        write_data_cell(ws, row_idx, 3, p.get("permit_number", ""))
        write_data_cell(ws, row_idx, 4, category)
        write_data_cell(ws, row_idx, 5, p.get("permit_authority", ""))
        write_data_cell(ws, row_idx, 6, p.get("issue_date", ""))
        write_data_cell(ws, row_idx, 7, p.get("expiry_date", ""))
        write_data_cell(ws, row_idx, 8, trade_text, alignment=ALIGN_LEFT)

        row_idx += 1

    # --- Freeze & Filter ---
    ws.freeze_panes = "A4"
    last_row = row_idx - 1
    if last_row >= 4:
        ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{last_row}"


# ---------------------------------------------------------------------------
# Sheet 3: 期限警告
# ---------------------------------------------------------------------------
def build_sheet_alerts(
    wb: openpyxl.Workbook,
    permits: list[dict[str, Any]],
    company_names: dict[str, str],
    permit_trades: dict[int, list[str]],
) -> None:
    ws = wb.create_sheet("期限警告")

    # --- Row 1: Title ---
    title_cell = ws.cell(row=1, column=1, value="有効期限アラート（2026年内）")
    title_cell.font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    # --- Row 2: empty (spacer) ---

    # --- Row 3: Headers ---
    headers = [
        "会社ID", "会社名", "許可番号", "許可区分",
        "許可者", "交付日", "有効期限", "許可業種",
    ]
    col_widths = [8, 30, 12, 8, 14, 12, 12, 50]
    apply_header_row(ws, 3, headers, col_widths)

    # --- Filter permits expiring within 365 days ---
    cutoff = TODAY + timedelta(days=365)
    alert_permits: list[dict[str, Any]] = []
    for p in permits:
        expiry_str = p.get("expiry_date")
        if not expiry_str or expiry_str == "UNCERTAIN":
            continue
        try:
            expiry_date = date.fromisoformat(expiry_str[:10])
        except (ValueError, TypeError):
            continue
        if expiry_date <= cutoff:
            alert_permits.append({**p, "_expiry_date": expiry_date})

    # 期限が近い順にソート
    alert_permits.sort(key=lambda x: x["_expiry_date"])

    # --- Data rows (row 4~) ---
    row_idx = 4
    for p in alert_permits:
        cid = p["company_id"]
        name = company_names.get(cid, cid)
        trades = permit_trades.get(p["permit_id"], [])
        trade_text = get_trade_short_names(trades)
        category = normalize_category(p.get("permit_category"))
        expiry_d: date = p["_expiry_date"]
        days_remaining = (expiry_d - TODAY).days

        # 行背景色
        if days_remaining <= 30:
            row_fill = FILL_RED
        elif days_remaining <= 90:
            row_fill = FILL_YELLOW
        else:
            row_fill = None

        write_data_cell(ws, row_idx, 1, cid, fill=row_fill)
        write_data_cell(ws, row_idx, 2, name, alignment=ALIGN_LEFT, fill=row_fill)
        write_data_cell(ws, row_idx, 3, p.get("permit_number", ""), fill=row_fill)
        write_data_cell(ws, row_idx, 4, category, fill=row_fill)
        write_data_cell(ws, row_idx, 5, p.get("permit_authority", ""), fill=row_fill)
        write_data_cell(ws, row_idx, 6, p.get("issue_date", ""), fill=row_fill)
        write_data_cell(ws, row_idx, 7, p.get("expiry_date", ""), fill=row_fill)
        write_data_cell(ws, row_idx, 8, trade_text, alignment=ALIGN_LEFT, fill=row_fill)

        row_idx += 1

    # --- Freeze & Filter ---
    ws.freeze_panes = "A4"
    last_row = row_idx - 1
    if last_row >= 4:
        ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{last_row}"


# ---------------------------------------------------------------------------
# Sheet 4: 許可業種一覧
# ---------------------------------------------------------------------------
def build_sheet_trade_master(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("許可業種一覧")

    # --- Row 1: Title ---
    title_cell = ws.cell(row=1, column=1, value="29業種 許可種類一覧")
    title_cell.font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    # --- Row 2: empty (spacer) ---

    # --- Row 3: Headers ---
    headers = ["No.", "業種コード", "業種名（正式）", "業種名（略称）"]
    col_widths = [5, 12, 30, 14]
    apply_header_row(ws, 3, headers, col_widths)

    # --- Data rows (row 4~) ---
    row_idx = 4
    for no, code, full_name, short_name in TRADE_29_MASTER:
        write_data_cell(ws, row_idx, 1, no)
        write_data_cell(ws, row_idx, 2, code)
        write_data_cell(ws, row_idx, 3, full_name, alignment=ALIGN_LEFT)
        write_data_cell(ws, row_idx, 4, short_name)
        row_idx += 1

    # --- Freeze ---
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# メイン
# ---------------------------------------------------------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description="建設業許可証管理台帳 Excel出力")
    parser.add_argument(
        "--output", "-o",
        type=str,
        default=None,
        help="出力ファイル名（デフォルト: permit_status_YYYYMMDD.xlsx）",
    )
    args = parser.parse_args()

    # 出力パス決定
    if args.output:
        output_path = Path(args.output)
        if not output_path.is_absolute():
            output_path = OUTPUT_DIR / output_path
    else:
        output_path = OUTPUT_DIR / f"permit_status_{TODAY.strftime('%Y%m%d')}.xlsx"

    output_path.parent.mkdir(parents=True, exist_ok=True)

    print("=" * 60)
    print(" 建設業許可証管理台帳 Excel出力")
    print("=" * 60)

    # --- DB読み取り ---
    conn = get_connection()
    companies = load_companies(conn)
    company_docs = load_company_docs(conn)
    permits = load_permits(conn)
    permit_trades_map = load_permit_trades(conn)
    company_names = load_company_names(conn)
    conn.close()

    print(f"  会社数: {len(companies)}")
    print(f"  許可証数: {len(permits)}")
    print(f"  書類受領会社: {len(company_docs)}")

    # --- Excel生成 ---
    wb = openpyxl.Workbook()

    # Sheet 1: 書類受領状況
    build_sheet_documents(wb, companies, company_docs)

    # Sheet 2: 許可証情報
    build_sheet_permits(wb, permits, company_names, permit_trades_map)

    # Sheet 3: 期限警告
    build_sheet_alerts(wb, permits, company_names, permit_trades_map)

    # Sheet 4: 許可業種一覧
    build_sheet_trade_master(wb)

    # --- 保存 ---
    wb.save(str(output_path))

    output_uri = str(output_path).replace("\\", "/")
    print(f"\n  出力: file:///{output_uri}")
    print(f"  サイズ: {output_path.stat().st_size / 1024:.1f} KB")
    print("=" * 60)


if __name__ == "__main__":
    main()
