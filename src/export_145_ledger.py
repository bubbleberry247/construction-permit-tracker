"""
export_145_ledger.py — 145社ベースのExcel台帳生成

145社Excelマスタ + permit_tracker.db からExcel台帳を生成する。
145社Excelの行順を維持し、DBにマッチしない行も「未受信」「未特定」で出力。

Usage:
    python src/export_145_ledger.py
"""
from __future__ import annotations

import re
import sqlite3
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

# ---------------------------------------------------------------------------
# パス定義
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).parent.parent
DB_PATH = PROJECT_ROOT / "data" / "permit_tracker.db"
EXCEL_145 = Path("C:/ProgramData/RK10/継続取引業者リスト_145社.xlsx")
OUTPUT_PATH = PROJECT_ROOT / "output" / "permit_status_20260329.xlsx"

# ---------------------------------------------------------------------------
# 書類種別マッピング（ヘッダ名 → DB doc_type_name）
# ---------------------------------------------------------------------------
# (ヘッダ名, DB doc_type_name, 必須フラグ)
DOC_COLUMNS: list[tuple[str, str, bool]] = [
    ("取引申請書", "取引申請書", True),
    ("建設業許可証", "建設業許可証", True),
    ("決算書", "決算書", True),
    ("会社案内", "会社案内", False),  # 任意（あれば）
    ("工事経歴書", "工事経歴書", True),
    ("取引先一覧表", "取引先一覧表", True),
    ("労安誓約書", "労働安全衛生誓約書", True),
    ("資格一覧", "資格略字一覧", True),
    ("労働者名簿", "労働者名簿", True),
]

# ---------------------------------------------------------------------------
# スタイル定数
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="0F3460")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=16)

STATUS_FILL_COMPLETE = PatternFill("solid", fgColor="E8F5E9")
STATUS_FILL_PARTIAL = PatternFill("solid", fgColor="FFF3E0")
STATUS_FILL_NONE = PatternFill("solid", fgColor="F5F5F5")
STATUS_FILL_UNKNOWN = PatternFill("solid", fgColor="FFF9C4")

FONT_GREY = Font(color="999999", size=10)
FONT_CHECK_OK = Font(bold=True, color="4CAF50", size=10)
FONT_CHECK_NG = Font(color="E0E0E0", size=10)
FONT_DEFAULT = Font(size=10)

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Sheet1列幅: A=5, B=8, C=30, D=14, E-M=8, N=35
SHEET1_COL_WIDTHS = [5, 8, 30, 14, 8, 8, 8, 8, 8, 8, 8, 8, 8, 35]


# ---------------------------------------------------------------------------
# 和暦変換ヘルパー
# ---------------------------------------------------------------------------
GANNEN_OFFSETS: dict[str, int] = {"令和": 2018, "R": 2018, "平成": 1988, "H": 1988}


def iso_to_wareki(iso_str: str | None) -> str:
    """ISO日付 (YYYY-MM-DD) → 令和表記 (R7.5.8)。変換不能時はそのまま返す。"""
    if not iso_str or iso_str == "UNCERTAIN":
        return iso_str or ""
    try:
        d = date.fromisoformat(iso_str[:10])
    except (ValueError, TypeError):
        return iso_str
    # 令和元年 = 2019
    reiwa_year = d.year - 2018
    if reiwa_year >= 1:
        return f"R{reiwa_year}.{d.month}.{d.day}"
    # 平成
    heisei_year = d.year - 1988
    if heisei_year >= 1:
        return f"H{heisei_year}.{d.month}.{d.day}"
    return iso_str


# ---------------------------------------------------------------------------
# DB読み取り
# ---------------------------------------------------------------------------
def get_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA foreign_keys = ON")
    conn.row_factory = sqlite3.Row
    return conn


def load_company_docs(conn: sqlite3.Connection) -> dict[str, set[str]]:
    """company_id → {doc_type_name, ...} のマップを返す。"""
    rows = conn.execute(
        """SELECT DISTINCT company_id, doc_type_name
           FROM pages
           WHERE company_id IS NOT NULL AND doc_type_name IS NOT NULL AND doc_type_name != ''"""
    ).fetchall()
    result: dict[str, set[str]] = {}
    for r in rows:
        result.setdefault(r["company_id"], set()).add(r["doc_type_name"])
    return result


def load_email_to_company(conn: sqlite3.Connection) -> dict[str, str]:
    """email → company_id のマップを返す。"""
    rows = conn.execute("SELECT email, company_id FROM company_emails").fetchall()
    result: dict[str, str] = {}
    for r in rows:
        email = r["email"].strip().lower()
        if email not in result:
            result[email] = r["company_id"]
    return result


def load_company_names(conn: sqlite3.Connection) -> dict[str, str]:
    """company_id → official_name"""
    rows = conn.execute("SELECT company_id, official_name FROM companies").fetchall()
    return {r["company_id"]: r["official_name"] for r in rows}


def load_companies_with_data(conn: sqlite3.Connection) -> set[str]:
    """何らかのデータ（inbound_messagesまたはpages）がある企業IDの集合。"""
    rows = conn.execute(
        """SELECT DISTINCT company_id FROM inbound_messages WHERE company_id IS NOT NULL
           UNION
           SELECT DISTINCT company_id FROM pages WHERE company_id IS NOT NULL"""
    ).fetchall()
    return {r[0] for r in rows}


def load_current_permits(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    """v_current_permits から全件取得。"""
    rows = conn.execute("SELECT * FROM v_current_permits ORDER BY company_id").fetchall()
    return [dict(r) for r in rows]


def load_expiry_2026(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    """2026年内に有効期限切れの許可証。"""
    rows = conn.execute(
        """SELECT p.permit_id, p.company_id, c.official_name, p.permit_number,
                  p.permit_category, p.permit_authority, p.issue_date, p.expiry_date,
                  GROUP_CONCAT(pt.trade_name, ',') AS trade_names
           FROM permits p
           JOIN companies c ON c.company_id = p.company_id
           LEFT JOIN permit_trades pt ON pt.permit_id = p.permit_id
           WHERE p.current_flag = 1
             AND p.expiry_date IS NOT NULL
             AND p.expiry_date LIKE '2026%'
           GROUP BY p.permit_id
           ORDER BY p.expiry_date"""
    ).fetchall()
    return [dict(r) for r in rows]


def load_permit_applications(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    """permit_applications テーブルの全件。"""
    rows = conn.execute(
        """SELECT pa.*, c.official_name
           FROM permit_applications pa
           JOIN companies c ON c.company_id = pa.company_id
           ORDER BY pa.application_date"""
    ).fetchall()
    return [dict(r) for r in rows]


# ---------------------------------------------------------------------------
# 145社Excel読み取り
# ---------------------------------------------------------------------------
def load_145_list() -> list[dict[str, str]]:
    """
    145社Excelを読み込み、行順を維持した辞書リストを返す。
    各辞書: {"company_name": str, "email": str, "is_email_only": bool}
    E列に改行で複数メアドがある場合は最初のメアドを使用。
    """
    wb = openpyxl.load_workbook(str(EXCEL_145), read_only=True)
    ws = wb.active
    result: list[dict[str, str]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        a_val = str(row[0] or "").strip()
        e_val = str(row[4] or "").strip() if len(row) > 4 else ""
        if not a_val:
            continue

        # E列に改行で複数メアドがある場合
        emails = [e.strip() for e in re.split(r"[\n\r]+", e_val) if e.strip()]

        is_email_only = "@" in a_val and "." in a_val and not any(
            c in a_val for c in "株有限会社"
        )

        result.append({
            "company_name": a_val,
            "emails": emails,
            "is_email_only": is_email_only,
        })

    wb.close()
    return result


# ---------------------------------------------------------------------------
# マッチング: 145社 → company_id
# ---------------------------------------------------------------------------
def match_145_to_db(
    list_145: list[dict[str, str]],
    email_to_cid: dict[str, str],
) -> list[dict[str, Any]]:
    """
    145社リストの各行にcompany_idを紐付ける。
    E列メアドでDB company_emailsを検索してcompany_id特定。
    """
    result: list[dict[str, Any]] = []
    for entry in list_145:
        cid = None
        for email in entry["emails"]:
            email_lower = email.lower()
            if email_lower in email_to_cid:
                cid = email_to_cid[email_lower]
                break
        result.append({
            **entry,
            "company_id": cid,
        })
    return result


# ---------------------------------------------------------------------------
# ステータス判定
# ---------------------------------------------------------------------------
def determine_status(
    company_id: str | None,
    is_email_only: bool,
    company_docs: dict[str, set[str]],
    companies_with_data: set[str],
) -> tuple[str, list[str]]:
    """
    ステータスと不足書類リストを返す。
    Returns: (status_label, missing_docs)
    """
    if is_email_only and company_id is None:
        return "未特定", []

    if company_id is None or company_id not in companies_with_data:
        return "未受信", []

    docs = company_docs.get(company_id, set())
    # 不足判定は必須書類のみ（会社案内は任意）
    missing = []
    for header_name, db_name, required in DOC_COLUMNS:
        if required and db_name not in docs:
            missing.append(header_name)

    if not missing:
        return "全揃い", []
    else:
        return f"不足あり ({len(missing)}種)", missing


# ---------------------------------------------------------------------------
# Sheet1: 書類受領状況
# ---------------------------------------------------------------------------
def write_sheet1(
    wb: openpyxl.Workbook,
    matched_list: list[dict[str, Any]],
    company_docs: dict[str, set[str]],
    company_names: dict[str, str],
    companies_with_data: set[str],
) -> None:
    ws = wb.active
    ws.title = "書類受領状況"

    # --- Row 1: タイトル ---
    ws["A1"] = "建設業許可証等 提出書類管理台帳"
    ws["A1"].font = TITLE_FONT

    # --- サマリー計算（重複company_id除外でカウント）---
    total = len(matched_list)
    count_received = 0
    count_complete = 0
    count_partial = 0
    count_not_received = 0

    statuses: list[tuple[str, list[str]]] = []
    summary_seen: set[str] = set()
    for entry in matched_list:
        status, missing = determine_status(
            entry["company_id"],
            entry["is_email_only"],
            company_docs,
            companies_with_data,
        )
        statuses.append((status, missing))

        # 重複company_idはサマリーカウントに含めない
        cid = entry["company_id"]
        if cid and cid in summary_seen:
            continue
        if cid:
            summary_seen.add(cid)

        if status == "全揃い":
            count_received += 1
            count_complete += 1
        elif status.startswith("不足あり"):
            count_received += 1
            count_partial += 1
        elif status in ("未受信", "未特定"):
            count_not_received += 1

    unique_total = len(summary_seen) + sum(1 for e in matched_list if not e["company_id"])

    # --- Row 2: サマリー ---
    ws["A2"] = (
        f"作成日: 2026-03-29 / 全{unique_total}社（送信先{total}メアド）"
        f"（受信{count_received}社・全揃{count_complete}社"
        f"・不足{count_partial}社・未受信{count_not_received}社）"
    )
    ws["A2"].font = Font(size=10)

    # --- Row 4: ヘッダ ---
    headers = ["No.", "会社ID", "会社名", "ステータス"]
    for header_name, _, _ in DOC_COLUMNS:
        headers.append(header_name)
    headers.append("不足書類")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    # --- Row 5+: データ行（重複company_idはスキップ）---
    seen_cids: set[str] = set()
    row_idx = 5
    for entry, (status, missing) in zip(matched_list, statuses):
        cid = entry["company_id"]

        # 同一company_idの重複行はスキップ
        if cid and cid in seen_cids:
            continue
        if cid:
            seen_cids.add(cid)

        if True:
            # 表示名: DB名 > Excelの会社名
            if cid and cid in company_names:
                display_name = company_names[cid]
            else:
                display_name = entry["company_name"]

            row_data = [
                row_idx - 4,           # No.
                cid or "",             # 会社ID
                display_name,          # 会社名
                status,                # ステータス
            ]

            # 書類列
            docs = company_docs.get(cid, set()) if cid else set()
            for _, db_name, _ in DOC_COLUMNS:
                row_data.append("○" if db_name in docs else "×")

            # 不足書類
            if missing:
                row_data.append("、".join(missing))
            elif status in ("未受信", "未特定"):
                row_data.append("")
            else:
                row_data.append("完備")

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER

            # ステータス列の背景色
            if status == "全揃い":
                cell.fill = STATUS_FILL_COMPLETE
            elif status.startswith("不足あり"):
                cell.fill = STATUS_FILL_PARTIAL
            elif status == "未受信":
                cell.fill = STATUS_FILL_NONE
                cell.font = FONT_GREY
            elif status == "未特定":
                cell.fill = STATUS_FILL_UNKNOWN

            # 書類列の文字色
            if col_idx >= 5 and col_idx <= 13:
                if val == "○":
                    cell.font = FONT_CHECK_OK
                    cell.alignment = ALIGN_CENTER
                elif val == "×":
                    cell.font = FONT_CHECK_NG
                    cell.alignment = ALIGN_CENTER
            elif col_idx == 1:  # No.
                cell.alignment = ALIGN_CENTER
            elif col_idx == 2:  # 会社ID
                cell.alignment = ALIGN_CENTER
            elif col_idx == 4:  # ステータス
                cell.alignment = ALIGN_CENTER

            # 未受信/未特定行のフォントを上書き（ただし書類列のフォントは個別に設定済み）
            if status == "未受信" and col_idx not in range(5, 14):
                cell.font = FONT_GREY

        row_idx += 1

    # --- 列幅 ---
    for col_idx, width in enumerate(SHEET1_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- freeze + filter ---
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:N{row_idx - 1}"


# ---------------------------------------------------------------------------
# Sheet2: 許可証情報
# ---------------------------------------------------------------------------
def write_sheet2(
    wb: openpyxl.Workbook,
    permits: list[dict[str, Any]],
    applications: list[dict[str, Any]],
) -> None:
    ws = wb.create_sheet("許可証情報")

    # --- Row 1: タイトル ---
    ws["A1"] = "建設業許可証 一覧"
    ws["A1"].font = Font(bold=True, size=14)

    # --- Row 3: ヘッダ ---
    headers = ["会社ID", "会社名", "許可番号", "許可区分", "許可者", "交付日", "有効期限", "許可業種"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    # --- Row 4+: permits ---
    row_idx = 4
    for p in permits:
        row_data = [
            p["company_id"],
            p["official_name"],
            p.get("permit_number", ""),
            p.get("permit_category", ""),
            p.get("permit_authority", ""),
            iso_to_wareki(p.get("issue_date")),
            iso_to_wareki(p.get("expiry_date")),
            p.get("trade_names", ""),
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.font = FONT_DEFAULT
            if col_idx in (1, 4):
                cell.alignment = ALIGN_CENTER
        row_idx += 1

    # --- permit_applications (申請中) ---
    for app in applications:
        row_data = [
            app["company_id"],
            app["official_name"],
            "(申請中)",
            "",
            "",
            "",
            "",
            app.get("applied_trades", ""),
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.font = Font(size=10, italic=True, color="FF6600")
            if col_idx in (1, 4):
                cell.alignment = ALIGN_CENTER
        row_idx += 1

    # --- 列幅 ---
    col_widths = [8, 30, 12, 8, 16, 14, 14, 50]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- freeze + filter ---
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:H{row_idx - 1}"


# ---------------------------------------------------------------------------
# Sheet3: 期限警告
# ---------------------------------------------------------------------------
def write_sheet3(
    wb: openpyxl.Workbook,
    expiry_2026: list[dict[str, Any]],
) -> None:
    ws = wb.create_sheet("期限警告")

    # --- Row 1: タイトル ---
    ws["A1"] = "有効期限 2026年内の許可証"
    ws["A1"].font = Font(bold=True, size=14)

    # --- Row 3: ヘッダ ---
    headers = ["会社ID", "会社名", "許可番号", "許可区分", "許可者", "交付日", "有効期限", "許可業種"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    # --- Row 4+: データ ---
    for row_idx, p in enumerate(expiry_2026, start=4):
        expiry_date_str = p.get("expiry_date", "")
        # 期限切れ判定
        try:
            expiry_date = date.fromisoformat(expiry_date_str[:10])
            is_expired = expiry_date < date(2026, 3, 29)
        except (ValueError, TypeError):
            is_expired = False

        row_data = [
            p["company_id"],
            p["official_name"],
            p.get("permit_number", ""),
            p.get("permit_category", ""),
            p.get("permit_authority", ""),
            iso_to_wareki(p.get("issue_date")),
            iso_to_wareki(expiry_date_str),
            p.get("trade_names", ""),
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if is_expired:
                cell.font = Font(size=10, color="FF0000", bold=True)
            else:
                cell.font = FONT_DEFAULT
            if col_idx in (1, 4):
                cell.alignment = ALIGN_CENTER

    # --- 列幅 ---
    col_widths = [8, 30, 12, 8, 16, 14, 14, 50]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    last_row = 3 + len(expiry_2026)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:H{last_row}"


# ---------------------------------------------------------------------------
# Sheet4: 申請中
# ---------------------------------------------------------------------------
def write_sheet4(
    wb: openpyxl.Workbook,
    applications: list[dict[str, Any]],
) -> None:
    ws = wb.create_sheet("申請中")

    # --- Row 1: タイトル ---
    ws["A1"] = "建設業許可 申請中一覧"
    ws["A1"].font = Font(bold=True, size=14)

    # --- Row 3: ヘッダ ---
    headers = ["会社ID", "会社名", "申請日", "申請業種", "ステータス", "備考"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    # --- Row 4+: データ ---
    for row_idx, app in enumerate(applications, start=4):
        row_data = [
            app["company_id"],
            app["official_name"],
            iso_to_wareki(app.get("application_date")),
            app.get("applied_trades", ""),
            app.get("status", ""),
            app.get("remarks", ""),
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.font = FONT_DEFAULT
            if col_idx == 1:
                cell.alignment = ALIGN_CENTER

    # --- 列幅 ---
    col_widths = [8, 30, 14, 40, 14, 50]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    last_row = 3 + len(applications)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:F{last_row}"


# ---------------------------------------------------------------------------
# メイン
# ---------------------------------------------------------------------------
def main() -> None:
    print("=" * 60)
    print(" 145社ベース Excel台帳生成")
    print("=" * 60)

    # --- DB読み取り ---
    conn = get_connection()
    company_docs = load_company_docs(conn)
    email_to_cid = load_email_to_company(conn)
    company_names = load_company_names(conn)
    companies_with_data = load_companies_with_data(conn)
    permits = load_current_permits(conn)
    expiry_2026 = load_expiry_2026(conn)
    applications = load_permit_applications(conn)

    print(f"  DB: {len(company_names)} 社, {len(email_to_cid)} メール, "
          f"{len(permits)} 許可証, {len(applications)} 申請中")

    # --- 145社Excel読み取り ---
    list_145 = load_145_list()
    print(f"  145社Excel: {len(list_145)} 行")

    # --- マッチング ---
    matched = match_145_to_db(list_145, email_to_cid)
    matched_count = sum(1 for m in matched if m["company_id"])
    unmatched_count = sum(1 for m in matched if not m["company_id"])
    email_only_count = sum(1 for m in matched if m["is_email_only"])
    print(f"  マッチ: {matched_count} 社, 未マッチ: {unmatched_count} 社, "
          f"メアドのみ: {email_only_count} 社")

    conn.close()

    # --- Excel出力 ---
    wb = openpyxl.Workbook()
    write_sheet1(wb, matched, company_docs, company_names, companies_with_data)
    write_sheet2(wb, permits, applications)
    write_sheet3(wb, expiry_2026)
    write_sheet4(wb, applications)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_PATH))
    print(f"\n  出力: file:///{str(OUTPUT_PATH).replace(chr(92), '/')}")
    print(f"  サイズ: {OUTPUT_PATH.stat().st_size / 1024:.1f} KB")
    print("=" * 60)


if __name__ == "__main__":
    main()
