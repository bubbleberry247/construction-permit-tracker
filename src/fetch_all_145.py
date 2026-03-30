"""
fetch_all_145.py — 145社全ての許可情報をMLITから取得して一覧CSV+Excel作成

DB上に許可番号がある会社はDB利用、ない会社は名前検索でMLITから取得。
"""
from __future__ import annotations

import csv
import random
import re
import sqlite3
import sys
import time
from datetime import date
from pathlib import Path

import requests
from bs4 import BeautifulSoup

sys.stdout.reconfigure(encoding="utf-8")

from verify_permits_full import (
    MLIT_SEARCH_URL,
    MLIT_TRADE_ABBREV,
    _post_with_retry,
    fetch_detail,
)

DB_PATH = Path(__file__).parent.parent / "data" / "permit_tracker.db"
OUTPUT_CSV = Path("C:/tmp/mlit_145_all.csv")
OUTPUT_XLSX = Path("C:/tmp/mlit_145_permit_list.xlsx")

CSV_HEADERS = [
    "company_id", "company_name", "permit_number", "authority",
    "category", "expiry_date", "days_remaining", "trades", "trades_count", "source",
]


def search_by_name(session: requests.Session, name: str) -> list[str]:
    """Search MLIT by company name, return list of sv_licenseNo."""
    # Clean name for search
    search_name = name.replace("株式会社", "").replace("有限会社", "").strip()
    if len(search_name) > 20:
        search_name = search_name[:20]

    params = {
        "CMD": "search", "caller": "KS", "rdoSelect": "2",
        "comNameKanaOnly": "", "comNameKanjiOnly": search_name,
        "rdoSelectJoken": "2",
        "licenseNoKbn": "", "licenseNoFrom": "", "licenseNoTo": "",
        "keyWord": "", "kenCode": "", "choice": "", "gyosyu": "", "gyosyuType": "",
        "sortValue": "", "rdoSelectSort": "1", "dispCount": "10", "dispPage": "1",
        "resultCount": "0", "pageCount": "0",
        "sv_rdoSelect": "", "sv_rdoSelectJoken": "", "sv_rdoSelectSort": "",
        "sv_kenCode": "", "sv_choice": "", "sv_gyosyu": "", "sv_gyosyuType": "",
        "sv_keyWord": "", "sv_sortValue": "", "sv_pageListNo1": "", "sv_pageListNo2": "",
        "sv_comNameKanaOnly": "", "sv_comNameKanjiOnly": "",
        "sv_licenseNoKbn": "", "sv_licenseNoFrom": "", "sv_licenseNoTo": "",
        "sv_licenseNo": "", "sv_dispCount": "0", "sv_dispPage": "0",
    }

    resp = _post_with_retry(session, MLIT_SEARCH_URL, params)
    html = resp.content.decode("cp932", errors="replace")
    soup = BeautifulSoup(html, "html.parser")

    hits = []
    for a_tag in soup.find_all("a", onclick=re.compile(r"js_ShowDetail")):
        m = re.search(r"js_ShowDetail\('(\d+)'\)", a_tag.get("onclick", ""))
        if m:
            hits.append(m.group(1))
    return hits


def safe_sleep():
    time.sleep(1.5 + random.uniform(0.3, 0.8))


def pref_from_sv(sv: str) -> str:
    """sv_licenseNo の先頭2桁から行政庁を推定."""
    code = sv[:2]
    pref_map = {
        "00": "国土交通大臣", "23": "愛知県知事", "21": "岐阜県知事",
        "06": "山形県知事", "22": "静岡県知事", "24": "三重県知事",
    }
    return pref_map.get(code, f"不明({code})")


def main():
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # All ACTIVE companies
    cur.execute("SELECT company_id, official_name FROM companies WHERE status = 'ACTIVE' ORDER BY official_name")
    companies = cur.fetchall()

    # Existing permit data from MLITPermits CSV (already scraped)
    mlit_csv = Path("C:/tmp/mlit_all_permits.csv")
    mlit_data = {}
    if mlit_csv.exists():
        with mlit_csv.open("r", encoding="utf-8-sig") as f:
            for r in csv.DictReader(f):
                if r.get("fetch_status") == "OK":
                    mlit_data[r["company_id"]] = r

    conn.close()

    # Session
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
        "Content-Type": "application/x-www-form-urlencoded",
    })

    print(f"[WARMUP]", end=" ", flush=True)
    resp = session.get(MLIT_SEARCH_URL, timeout=15)
    print(f"OK ({resp.status_code})")

    results: list[dict] = []
    total = len(companies)
    searched = 0
    found_new = 0

    for i, c in enumerate(companies):
        cid = c["company_id"]
        name = c["official_name"]

        # Already have MLIT data?
        if cid in mlit_data:
            m = mlit_data[cid]
            days = int(m.get("days_remaining", 0)) if m.get("days_remaining") else 0
            results.append({
                "company_id": cid, "company_name": name,
                "permit_number": m.get("permit_number", ""),
                "authority": m.get("authority", ""),
                "category": m.get("category", ""),
                "expiry_date": m.get("expiry_date", ""),
                "days_remaining": days,
                "trades": m.get("trades_ippan", ""),
                "trades_count": m.get("trades_count", ""),
                "source": "MLIT_CACHED",
            })
            continue

        # Name search
        print(f"  [{i+1}/{total}] {name}", end=" ... ", flush=True)
        searched += 1

        try:
            hits = search_by_name(session, name)

            if hits:
                sv = hits[0]
                safe_sleep()
                detail = fetch_detail(session, sv)

                if detail.found:
                    all_trades = detail.api_trades_ippan + detail.api_trades_tokutei
                    cat = "般" if detail.api_trades_ippan and not detail.api_trades_tokutei else \
                          "特" if detail.api_trades_tokutei and not detail.api_trades_ippan else \
                          "般/特" if detail.api_trades_ippan and detail.api_trades_tokutei else "-"
                    exp = detail.api_expiry_to
                    days = (date.fromisoformat(exp) - date.today()).days if exp else 0
                    results.append({
                        "company_id": cid, "company_name": name,
                        "permit_number": sv[-6:].lstrip("0") or "0",
                        "authority": pref_from_sv(sv),
                        "category": cat,
                        "expiry_date": exp,
                        "days_remaining": days,
                        "trades": "|".join(all_trades),
                        "trades_count": len(all_trades),
                        "source": "MLIT_NAME",
                    })
                    found_new += 1
                    print(f"FOUND ({detail.api_name}, exp={exp})")
                else:
                    results.append({
                        "company_id": cid, "company_name": name,
                        "permit_number": "", "authority": "", "category": "",
                        "expiry_date": "", "days_remaining": "",
                        "trades": "", "trades_count": 0, "source": "NOT_FOUND",
                    })
                    print("DETAIL_FAIL")
            else:
                results.append({
                    "company_id": cid, "company_name": name,
                    "permit_number": "", "authority": "", "category": "",
                    "expiry_date": "", "days_remaining": "",
                    "trades": "", "trades_count": 0, "source": "NOT_FOUND",
                })
                print("NOT_FOUND")

        except Exception as e:
            results.append({
                "company_id": cid, "company_name": name,
                "permit_number": "", "authority": "", "category": "",
                "expiry_date": "", "days_remaining": "",
                "trades": "", "trades_count": 0, "source": f"ERROR:{e}",
            })
            print(f"ERROR: {e}")

        safe_sleep()

    # Write CSV
    with OUTPUT_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_HEADERS)
        writer.writeheader()
        for r in results:
            writer.writerow(r)

    # Write Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "145社許可一覧"

        thin = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(name="Meiryo", bold=True, color="FFFFFF", size=10)
        cell_font = Font(name="Meiryo", size=9)
        red_fill = PatternFill("solid", fgColor="FFC7CE")
        yellow_fill = PatternFill("solid", fgColor="FFEB9C")
        green_fill = PatternFill("solid", fgColor="C6EFCE")

        headers = ["会社名", "許可番号", "許可行政庁", "般/特", "有効期限", "残日数", "業種", "業種数", "取得元"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")

        results.sort(key=lambda x: (x.get("days_remaining") if isinstance(x.get("days_remaining"), int) and x.get("days_remaining") else 99999))

        for row_idx, r in enumerate(results, 2):
            vals = [
                r["company_name"], r["permit_number"], r["authority"], r["category"],
                r["expiry_date"], r["days_remaining"],
                r["trades"].replace("|", ", "), r["trades_count"], r["source"],
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col, value=v)
                cell.font = cell_font
                cell.border = thin
                if col == 6 and isinstance(v, int):
                    if v <= 0:
                        cell.fill = red_fill
                    elif v <= 90:
                        cell.fill = yellow_fill
                    elif v <= 150:
                        cell.fill = green_fill

        # Column widths
        widths = [30, 10, 14, 6, 12, 8, 50, 6, 14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        ws.freeze_panes = "A2"
        wb.save(str(OUTPUT_XLSX))
        print(f"\nExcel: {OUTPUT_XLSX}")
    except Exception as e:
        print(f"\nExcel error: {e}")

    # Summary
    sources = {}
    for r in results:
        s = r["source"]
        sources[s] = sources.get(s, 0) + 1

    print(f"\n{'='*60}")
    print(f"  145社 許可情報取得 完了")
    print(f"{'='*60}")
    print(f"  Total: {len(results)}")
    print(f"  Name searched: {searched}")
    print(f"  New found: {found_new}")
    for s, cnt in sorted(sources.items()):
        print(f"    {s}: {cnt}")
    print(f"  CSV: {OUTPUT_CSV}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
