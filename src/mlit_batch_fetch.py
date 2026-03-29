"""
mlit_batch_fetch.py — MLIT全社スクレイピング + 分析 + Excel + DB更新

全7フェーズを自動実行:
  Phase 1: 全社スクレイピング (permits テーブル)
  Phase 2: DB未登録社の追加スクレイピング (companies テーブル)
  Phase 3: DB vs MLIT 照合レポート
  Phase 4: 期限アラートレポート
  Phase 5: 業種マトリクスExcel生成
  Phase 6: DB更新
  Phase 7: サマリーレポート

Usage:
    python src/mlit_batch_fetch.py              # 全フェーズ実行
    python src/mlit_batch_fetch.py --dry-run    # APIアクセスなし
    python src/mlit_batch_fetch.py --resume     # 中断再開
"""

from __future__ import annotations

import argparse
import csv
import random
import sqlite3
import sys
import time
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

# Reuse from verify_permits_full.py
from verify_permits_full import (
    DB_PATH,
    FORMAL_TO_ABBREV,
    MLIT_DETAIL_URL,
    MLIT_SEARCH_URL,
    MLIT_TRADE_ABBREV,
    REQUEST_INTERVAL_SEC,
    ApiDetail,
    _post_with_retry,
    fetch_detail,
    get_license_no_kbn,
    get_pref_code,
    normalize_permit_number,
    search_permit,
    wareki_period_to_iso,
)

try:
    import requests
except ImportError:
    print("ERROR: requests not installed")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

OUTPUT_DIR = Path("C:/tmp")
CSV_ALL = OUTPUT_DIR / "mlit_all_permits.csv"
CSV_COMPARISON = OUTPUT_DIR / "mlit_comparison_report.csv"
CSV_ALERTS = OUTPUT_DIR / "mlit_expiry_alerts.csv"
XLSX_REGISTRY = OUTPUT_DIR / "mlit_permit_registry.xlsx"
CSV_DB_UPDATE_LOG = OUTPUT_DIR / "mlit_db_update_log.csv"
TXT_SUMMARY = OUTPUT_DIR / "mlit_batch_summary.txt"

SAFE_INTERVAL_BASE = 2.0
JITTER_MIN = 0.3
JITTER_MAX = 0.8

CSV_HEADERS = [
    "company_id", "company_name", "permit_number", "authority",
    "category", "expiry_date", "expiry_wareki", "days_remaining",
    "trades_ippan", "trades_tokutei", "trades_count",
    "fetch_status", "error",
]


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class FetchRow:
    company_id: str
    company_name: str
    permit_number: str
    authority: str
    category: str = ""
    expiry_date: str = ""
    expiry_wareki: str = ""
    days_remaining: int = 0
    trades_ippan: str = ""
    trades_tokutei: str = ""
    trades_count: int = 0
    fetch_status: str = ""
    error: str = ""


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def safe_sleep() -> None:
    """Rate-limited sleep with jitter."""
    delay = SAFE_INTERVAL_BASE + random.uniform(JITTER_MIN, JITTER_MAX)
    time.sleep(delay)


def calc_days_remaining(expiry_iso: str) -> int:
    """Calculate days until expiry. Negative = expired."""
    if not expiry_iso:
        return 0
    try:
        exp = date.fromisoformat(expiry_iso)
        return (exp - date.today()).days
    except ValueError:
        return 0


def load_resume_keys(csv_path: Path) -> set[str]:
    """Load already-processed permit_number+authority keys from CSV."""
    keys: set[str] = set()
    if not csv_path.exists():
        return keys
    with csv_path.open("r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            key = f"{row.get('authority', '')}|{row.get('permit_number', '')}"
            keys.add(key)
    return keys


# ---------------------------------------------------------------------------
# Phase 1 & 2: Scrape
# ---------------------------------------------------------------------------

def fetch_one_company(
    session: requests.Session,
    company_id: str,
    company_name: str,
    permit_number: str,
    authority: str,
) -> FetchRow:
    """Fetch one company from MLIT and return a FetchRow."""
    row = FetchRow(
        company_id=company_id,
        company_name=company_name,
        permit_number=permit_number,
        authority=authority,
    )

    license_no_kbn = get_license_no_kbn(authority)
    pref_code = get_pref_code(authority)

    try:
        sv = search_permit(session, license_no_kbn, permit_number, pref_code)
    except Exception as exc:
        row.fetch_status = "ERROR"
        row.error = f"search: {exc}"
        return row

    safe_sleep()

    if not sv:
        row.fetch_status = "NOT_FOUND"
        row.error = "Search returned no results"
        return row

    try:
        detail = fetch_detail(session, sv)
    except Exception as exc:
        row.fetch_status = "ERROR"
        row.error = f"detail: {exc}"
        return row

    if not detail.found:
        row.fetch_status = "NOT_FOUND"
        row.error = detail.error or "Detail page empty"
        return row

    # Determine category: if both ippan and tokutei exist, take the one with more trades
    if detail.api_trades_tokutei and not detail.api_trades_ippan:
        row.category = "特"
    elif detail.api_trades_ippan and not detail.api_trades_tokutei:
        row.category = "般"
    elif detail.api_trades_ippan and detail.api_trades_tokutei:
        row.category = "般/特"
    else:
        row.category = "-"

    row.expiry_date = detail.api_expiry_to
    row.expiry_wareki = detail.api_expiry_wareki
    row.days_remaining = calc_days_remaining(detail.api_expiry_to)
    row.trades_ippan = "|".join(detail.api_trades_ippan)
    row.trades_tokutei = "|".join(detail.api_trades_tokutei)
    row.trades_count = len(detail.api_trades_ippan) + len(detail.api_trades_tokutei)
    row.fetch_status = "OK"

    return row


def write_csv_row(f, writer: csv.DictWriter, row: FetchRow) -> None:
    """Write one row and flush immediately."""
    writer.writerow({
        "company_id": row.company_id,
        "company_name": row.company_name,
        "permit_number": row.permit_number,
        "authority": row.authority,
        "category": row.category,
        "expiry_date": row.expiry_date,
        "expiry_wareki": row.expiry_wareki,
        "days_remaining": row.days_remaining,
        "trades_ippan": row.trades_ippan,
        "trades_tokutei": row.trades_tokutei,
        "trades_count": row.trades_count,
        "fetch_status": row.fetch_status,
        "error": row.error,
    })
    f.flush()


def run_scrape(
    session: requests.Session,
    resume_keys: set[str],
    dry_run: bool = False,
) -> list[FetchRow]:
    """Phase 1 + 2: Scrape all companies."""

    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Phase 1: permits table
    cur.execute("""
        SELECT DISTINCT p.company_id, c.official_name, p.permit_number, p.permit_authority
        FROM permits p
        JOIN companies c ON p.company_id = c.company_id
        WHERE p.permit_number IS NOT NULL AND p.permit_number != ''
          AND p.permit_authority IS NOT NULL AND p.permit_authority != ''
        ORDER BY p.company_id
    """)
    phase1_rows = cur.fetchall()

    # Phase 2: companies table (have permit_number but no permit record)
    cur.execute("""
        SELECT c.company_id, c.official_name, c.permit_number, c.permit_authority
        FROM companies c
        WHERE c.permit_number IS NOT NULL AND c.permit_number != ''
          AND c.permit_authority IS NOT NULL AND c.permit_authority != ''
          AND NOT EXISTS (
              SELECT 1 FROM permits p
              WHERE p.company_id = c.company_id
                AND p.permit_number = c.permit_number
          )
        ORDER BY c.company_id
    """)
    phase2_rows = cur.fetchall()
    conn.close()

    all_targets = []
    seen_keys: set[str] = set()

    for r in phase1_rows:
        key = f"{r['permit_authority']}|{r['permit_number']}"
        if key not in seen_keys:
            seen_keys.add(key)
            all_targets.append(("P1", r["company_id"], r["official_name"], r["permit_number"], r["permit_authority"]))

    for r in phase2_rows:
        key = f"{r['permit_authority']}|{r['permit_number']}"
        if key not in seen_keys:
            seen_keys.add(key)
            all_targets.append(("P2", r["company_id"], r["official_name"], r["permit_number"], r["permit_authority"]))

    total = len(all_targets)
    skipped_resume = 0

    print(f"\n{'='*60}")
    print(f"  Phase 1+2: MLIT Scraping")
    print(f"  Total targets: {total} (Phase1: {len(phase1_rows)}, Phase2: {len(phase2_rows)} unique additions)")
    print(f"  Resume keys loaded: {len(resume_keys)}")
    print(f"{'='*60}")

    if dry_run:
        print("\n[DRY-RUN] Skipping API access")
        for phase, cid, name, num, auth in all_targets:
            key = f"{auth}|{num}"
            skip = "[SKIP-RESUME]" if key in resume_keys else ""
            print(f"  [{phase}] {cid} {name} | {auth} {num} {skip}")
        return []

    # Server warmup
    print("\n[WARMUP] Checking server ...", end=" ", flush=True)
    try:
        resp = session.get(MLIT_SEARCH_URL, timeout=15)
        if resp.status_code < 500:
            print(f"OK (HTTP {resp.status_code})")
        else:
            print(f"HTTP {resp.status_code} - proceeding anyway")
    except Exception as e:
        print(f"Warning: {e} - proceeding anyway")

    # Open CSV for incremental write
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    is_new_file = not CSV_ALL.exists() or not resume_keys
    mode = "a" if resume_keys and CSV_ALL.exists() else "w"

    f = CSV_ALL.open(mode, encoding="utf-8-sig", newline="")
    writer = csv.DictWriter(f, fieldnames=CSV_HEADERS)
    if is_new_file or mode == "w":
        writer.writeheader()

    results: list[FetchRow] = []
    success = 0
    fail = 0
    consecutive_403 = 0

    for i, (phase, cid, name, num, auth) in enumerate(all_targets):
        key = f"{auth}|{num}"

        if key in resume_keys:
            skipped_resume += 1
            continue

        print(
            f"  [{i+1}/{total}] [{phase}] {name} ({auth} {num})",
            end=" ... ",
            flush=True,
        )

        row = fetch_one_company(session, cid, name, num, auth)
        results.append(row)
        write_csv_row(f, writer, row)

        if row.fetch_status == "OK":
            success += 1
            consecutive_403 = 0
            print(f"OK ({row.trades_count} trades, exp={row.expiry_date}, {row.days_remaining}d)")
        elif row.fetch_status == "NOT_FOUND":
            fail += 1
            consecutive_403 = 0
            print(f"NOT_FOUND")
        else:
            fail += 1
            print(f"ERROR: {row.error}")
            if "403" in row.error:
                consecutive_403 += 1
                if consecutive_403 >= 3:
                    print("\n[STOP] 3 consecutive 403 errors - possible IP block. Stopping.")
                    break

        # Rate limit between companies
        if i < total - 1:
            safe_sleep()

    f.close()

    print(f"\n[Phase 1+2 Complete] success={success} fail={fail} skipped_resume={skipped_resume}")
    print(f"  CSV: {CSV_ALL}")

    return results


# ---------------------------------------------------------------------------
# Phase 3: DB vs MLIT comparison
# ---------------------------------------------------------------------------

def run_comparison(fetch_results: list[FetchRow]) -> list[dict]:
    """Compare MLIT data with DB data."""
    print(f"\n{'='*60}")
    print(f"  Phase 3: DB vs MLIT Comparison")
    print(f"{'='*60}")

    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    cur.execute("""
        SELECT p.permit_id, p.company_id, c.official_name, p.permit_number,
               p.permit_category, p.permit_authority, p.expiry_date,
               GROUP_CONCAT(pt.trade_name, ',') as trades
        FROM permits p
        JOIN companies c ON p.company_id = c.company_id
        LEFT JOIN permit_trades pt ON p.permit_id = pt.permit_id
        GROUP BY p.permit_id
    """)
    db_permits = {f"{r['permit_authority']}|{r['permit_number']}": dict(r) for r in cur.fetchall()}
    conn.close()

    # Build lookup from fetch results
    mlit_lookup = {}
    for fr in fetch_results:
        if fr.fetch_status == "OK":
            mlit_lookup[f"{fr.authority}|{fr.permit_number}"] = fr

    rows = []
    for key, db in db_permits.items():
        mlit = mlit_lookup.get(key)
        if not mlit:
            rows.append({
                "company_name": db["official_name"],
                "permit_number": db["permit_number"],
                "authority": db["permit_authority"],
                "field": "ALL",
                "db_value": "-",
                "mlit_value": "NOT_FETCHED",
                "match": "-",
            })
            continue

        # Expiry comparison
        db_exp = db["expiry_date"] or ""
        mlit_exp = mlit.expiry_date or ""
        exp_match = "OK" if db_exp == mlit_exp else "NG"
        if exp_match == "NG":
            rows.append({
                "company_name": db["official_name"],
                "permit_number": db["permit_number"],
                "authority": db["permit_authority"],
                "field": "expiry_date",
                "db_value": db_exp,
                "mlit_value": mlit_exp,
                "match": "NG",
            })

        # Trade comparison
        db_trades_raw = db["trades"] or ""
        db_trades_set = set()
        for t in db_trades_raw.split(","):
            t = t.strip()
            if t:
                norm = FORMAL_TO_ABBREV.get(t, t.replace("工事業", "").replace("工事", ""))
                db_trades_set.add(norm)

        mlit_trades_set = set()
        if mlit.trades_ippan:
            mlit_trades_set.update(mlit.trades_ippan.split("|"))
        if mlit.trades_tokutei:
            mlit_trades_set.update(mlit.trades_tokutei.split("|"))

        if db_trades_set != mlit_trades_set:
            missing = db_trades_set - mlit_trades_set
            extra = mlit_trades_set - db_trades_set
            diff_parts = []
            if missing:
                diff_parts.append(f"DB only: {','.join(sorted(missing))}")
            if extra:
                diff_parts.append(f"MLIT only: {','.join(sorted(extra))}")
            rows.append({
                "company_name": db["official_name"],
                "permit_number": db["permit_number"],
                "authority": db["permit_authority"],
                "field": "trades",
                "db_value": ",".join(sorted(db_trades_set)),
                "mlit_value": ",".join(sorted(mlit_trades_set)),
                "match": "NG" if missing else "EXTRA",
            })

    # Write CSV
    comp_headers = ["company_name", "permit_number", "authority", "field", "db_value", "mlit_value", "match"]
    with CSV_COMPARISON.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=comp_headers)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)

    ng_count = sum(1 for r in rows if r["match"] == "NG")
    print(f"  Differences found: {len(rows)} ({ng_count} NG)")
    print(f"  CSV: {CSV_COMPARISON}")

    return rows


# ---------------------------------------------------------------------------
# Phase 4: Expiry alerts
# ---------------------------------------------------------------------------

def run_expiry_alerts(fetch_results: list[FetchRow]) -> list[dict]:
    """Generate expiry alert report."""
    print(f"\n{'='*60}")
    print(f"  Phase 4: Expiry Alerts (120 days)")
    print(f"{'='*60}")

    alerts = []
    for fr in fetch_results:
        if fr.fetch_status != "OK":
            continue
        if fr.days_remaining <= 120:
            stage = "EXPIRED" if fr.days_remaining <= 0 else \
                    "30_DAYS" if fr.days_remaining <= 30 else \
                    "60_DAYS" if fr.days_remaining <= 60 else \
                    "90_DAYS" if fr.days_remaining <= 90 else \
                    "120_DAYS"
            all_trades = []
            if fr.trades_ippan:
                all_trades.extend(fr.trades_ippan.split("|"))
            if fr.trades_tokutei:
                all_trades.extend(fr.trades_tokutei.split("|"))

            alerts.append({
                "stage": stage,
                "company_name": fr.company_name,
                "company_id": fr.company_id,
                "permit_number": fr.permit_number,
                "authority": fr.authority,
                "category": fr.category,
                "expiry_date": fr.expiry_date,
                "days_remaining": fr.days_remaining,
                "trades": "|".join(all_trades),
            })

    alerts.sort(key=lambda x: x["days_remaining"])

    alert_headers = ["stage", "company_name", "company_id", "permit_number", "authority", "category", "expiry_date", "days_remaining", "trades"]
    with CSV_ALERTS.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=alert_headers)
        writer.writeheader()
        for a in alerts:
            writer.writerow(a)

    # Summary by stage
    stage_counts: dict[str, int] = {}
    for a in alerts:
        stage_counts[a["stage"]] = stage_counts.get(a["stage"], 0) + 1

    print(f"  Total alerts: {len(alerts)}")
    for stage in ["EXPIRED", "30_DAYS", "60_DAYS", "90_DAYS", "120_DAYS"]:
        if stage in stage_counts:
            print(f"    {stage}: {stage_counts[stage]}")
    print(f"  CSV: {CSV_ALERTS}")

    return alerts


# ---------------------------------------------------------------------------
# Phase 5: Excel
# ---------------------------------------------------------------------------

def run_excel(fetch_results: list[FetchRow], comparison: list[dict], alerts: list[dict]) -> None:
    """Generate Excel registry with 4 sheets."""
    print(f"\n{'='*60}")
    print(f"  Phase 5: Excel Generation")
    print(f"{'='*60}")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        print("  [SKIP] openpyxl not installed. Run: pip install openpyxl")
        return

    wb = Workbook()
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Meiryo", bold=True, color="FFFFFF", size=10)
    cell_font = Font(name="Meiryo", size=9)
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    ok_results = [r for r in fetch_results if r.fetch_status == "OK"]
    ok_results.sort(key=lambda x: x.days_remaining)

    # --- Sheet 1: 全社一覧 ---
    ws1 = wb.active
    ws1.title = "全社一覧"
    s1_headers = ["会社名", "許可番号", "許可行政庁", "区分", "有効期限", "残日数", "業種数", "業種（般）", "業種（特）"]
    for col, h in enumerate(s1_headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row_idx, r in enumerate(ok_results, 2):
        vals = [
            r.company_name, r.permit_number, r.authority, r.category,
            r.expiry_date, r.days_remaining, r.trades_count,
            r.trades_ippan.replace("|", ", "),
            r.trades_tokutei.replace("|", ", "),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws1.cell(row=row_idx, column=col, value=v)
            cell.font = cell_font
            cell.border = thin_border
            # Color by days_remaining
            if col == 6 and isinstance(v, int):
                if v <= 0:
                    cell.fill = red_fill
                elif v <= 90:
                    cell.fill = yellow_fill
                elif v <= 120:
                    cell.fill = green_fill

    # Auto-width
    for col_idx in range(1, len(s1_headers) + 1):
        max_len = max(
            (len(str(ws1.cell(row=r, column=col_idx).value or "")) for r in range(1, len(ok_results) + 2)),
            default=10,
        )
        ws1.column_dimensions[ws1.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

    # --- Sheet 2: 業種マトリクス ---
    ws2 = wb.create_sheet("業種マトリクス")
    s2_headers = ["会社名", "区分"] + MLIT_TRADE_ABBREV
    for col, h in enumerate(s2_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", text_rotation=90 if col > 2 else 0)

    for row_idx, r in enumerate(ok_results, 2):
        ws2.cell(row=row_idx, column=1, value=r.company_name).font = cell_font
        ws2.cell(row=row_idx, column=1).border = thin_border
        ws2.cell(row=row_idx, column=2, value=r.category).font = cell_font
        ws2.cell(row=row_idx, column=2).border = thin_border

        ippan_set = set(r.trades_ippan.split("|")) if r.trades_ippan else set()
        tokutei_set = set(r.trades_tokutei.split("|")) if r.trades_tokutei else set()

        for t_idx, trade in enumerate(MLIT_TRADE_ABBREV):
            col = t_idx + 3
            cell = ws2.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            cell.font = cell_font
            if trade in tokutei_set:
                cell.value = "●"
            elif trade in ippan_set:
                cell.value = "○"
            else:
                cell.value = ""

    # Column widths for trade matrix
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 6
    for t_idx in range(len(MLIT_TRADE_ABBREV)):
        col_letter = ws2.cell(row=1, column=t_idx + 3).column_letter
        ws2.column_dimensions[col_letter].width = 4

    # --- Sheet 3: 期限アラート ---
    ws3 = wb.create_sheet("期限アラート")
    s3_headers = ["ステージ", "会社名", "許可番号", "許可行政庁", "区分", "有効期限", "残日数", "業種"]
    for col, h in enumerate(s3_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row_idx, a in enumerate(alerts, 2):
        vals = [
            a["stage"], a["company_name"], a["permit_number"], a["authority"],
            a["category"], a["expiry_date"], a["days_remaining"],
            a["trades"].replace("|", ", "),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws3.cell(row=row_idx, column=col, value=v)
            cell.font = cell_font
            cell.border = thin_border

        # Row color by stage
        stage = a["stage"]
        fill = red_fill if stage in ("EXPIRED", "30_DAYS") else \
               yellow_fill if stage in ("60_DAYS", "90_DAYS") else \
               green_fill
        for col in range(1, len(s3_headers) + 1):
            ws3.cell(row=row_idx, column=col).fill = fill

    # --- Sheet 4: DB vs MLIT差異 ---
    ws4 = wb.create_sheet("DB差異")
    ng_rows = [r for r in comparison if r["match"] != "OK"]
    s4_headers = ["会社名", "許可番号", "項目", "DB値", "MLIT値", "判定"]
    for col, h in enumerate(s4_headers, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row_idx, r in enumerate(ng_rows, 2):
        vals = [r["company_name"], r["permit_number"], r["field"], r["db_value"], r["mlit_value"], r["match"]]
        for col, v in enumerate(vals, 1):
            cell = ws4.cell(row=row_idx, column=col, value=v)
            cell.font = cell_font
            cell.border = thin_border
            if r["match"] == "NG":
                cell.fill = red_fill

    if not ng_rows:
        ws4.cell(row=2, column=1, value="差異なし").font = Font(name="Meiryo", size=11, color="008000")

    # Save
    wb.save(str(XLSX_REGISTRY))
    print(f"  Excel saved: {XLSX_REGISTRY}")
    print(f"    Sheet1 全社一覧: {len(ok_results)} rows")
    print(f"    Sheet2 業種マトリクス: {len(ok_results)} rows x {len(MLIT_TRADE_ABBREV)} trades")
    print(f"    Sheet3 期限アラート: {len(alerts)} rows")
    print(f"    Sheet4 DB差異: {len(ng_rows)} rows")


# ---------------------------------------------------------------------------
# Phase 6: DB update
# ---------------------------------------------------------------------------

def run_db_update(fetch_results: list[FetchRow]) -> list[dict]:
    """Update DB with MLIT confirmed data."""
    print(f"\n{'='*60}")
    print(f"  Phase 6: DB Update")
    print(f"{'='*60}")

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    updates: list[dict] = []

    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    for fr in fetch_results:
        if fr.fetch_status != "OK":
            continue

        # Update companies.mlit_status + last_confirmed_at
        cur.execute(
            "UPDATE companies SET mlit_status = 'CONFIRMED', last_confirmed_at = ? WHERE company_id = ?",
            (now_str, fr.company_id),
        )

        # Check and update permit expiry_date if different
        cur.execute("""
            SELECT permit_id, expiry_date FROM permits
            WHERE company_id = ? AND permit_number = ?
        """, (fr.company_id, fr.permit_number))
        permit_row = cur.fetchone()

        if permit_row:
            pid = permit_row["permit_id"]
            old_exp = permit_row["expiry_date"] or ""
            new_exp = fr.expiry_date or ""

            if old_exp != new_exp and new_exp:
                cur.execute(
                    "UPDATE permits SET expiry_date = ?, updated_at = ? WHERE permit_id = ?",
                    (new_exp, now_str, pid),
                )
                updates.append({
                    "company_name": fr.company_name,
                    "permit_number": fr.permit_number,
                    "field": "expiry_date",
                    "old_value": old_exp,
                    "new_value": new_exp,
                    "action": "UPDATED",
                })

            # Update permit_trades: get current, compare, add missing
            cur.execute("SELECT trade_name FROM permit_trades WHERE permit_id = ?", (pid,))
            db_trades = {FORMAL_TO_ABBREV.get(r["trade_name"], r["trade_name"].replace("工事業", "").replace("工事", "")) for r in cur.fetchall()}

            mlit_trades: set[str] = set()
            if fr.trades_ippan:
                mlit_trades.update(fr.trades_ippan.split("|"))
            if fr.trades_tokutei:
                mlit_trades.update(fr.trades_tokutei.split("|"))

            missing_in_db = mlit_trades - db_trades
            if missing_in_db:
                for trade in missing_in_db:
                    # Store with 工事業 suffix for consistency
                    formal_name = trade + "工事業" if not trade.endswith("工事業") else trade
                    cur.execute(
                        "INSERT INTO permit_trades (permit_id, trade_name) VALUES (?, ?)",
                        (pid, formal_name),
                    )
                updates.append({
                    "company_name": fr.company_name,
                    "permit_number": fr.permit_number,
                    "field": "trades_added",
                    "old_value": ",".join(sorted(db_trades)),
                    "new_value": ",".join(sorted(missing_in_db)),
                    "action": "ADDED",
                })

    conn.commit()
    conn.close()

    # Write update log
    log_headers = ["company_name", "permit_number", "field", "old_value", "new_value", "action"]
    with CSV_DB_UPDATE_LOG.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=log_headers)
        writer.writeheader()
        for u in updates:
            writer.writerow(u)

    confirmed = sum(1 for fr in fetch_results if fr.fetch_status == "OK")
    print(f"  Companies confirmed: {confirmed}")
    print(f"  DB changes: {len(updates)}")
    print(f"  Update log: {CSV_DB_UPDATE_LOG}")

    return updates


# ---------------------------------------------------------------------------
# Phase 7: Summary
# ---------------------------------------------------------------------------

def run_summary(
    fetch_results: list[FetchRow],
    comparison: list[dict],
    alerts: list[dict],
    db_updates: list[dict],
) -> None:
    """Generate summary report."""
    print(f"\n{'='*60}")
    print(f"  Phase 7: Summary Report")
    print(f"{'='*60}")

    total = len(fetch_results)
    ok = sum(1 for r in fetch_results if r.fetch_status == "OK")
    not_found = sum(1 for r in fetch_results if r.fetch_status == "NOT_FOUND")
    errors = sum(1 for r in fetch_results if r.fetch_status == "ERROR")

    # Trade stats
    all_trades: dict[str, int] = {}
    for fr in fetch_results:
        if fr.fetch_status == "OK":
            for t in (fr.trades_ippan + "|" + fr.trades_tokutei).split("|"):
                if t:
                    all_trades[t] = all_trades.get(t, 0) + 1

    comp_ng = sum(1 for r in comparison if r["match"] == "NG")

    lines = [
        "=" * 60,
        "  MLIT 全社スクレイピング サマリーレポート",
        f"  実行日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "=" * 60,
        "",
        "--- スクレイピング結果 ---",
        f"  対象:      {total} 社",
        f"  成功:      {ok} 社 ({ok/total*100:.0f}%)" if total > 0 else "  成功: 0",
        f"  NOT_FOUND: {not_found} 社",
        f"  ERROR:     {errors} 社",
        "",
        "--- DB vs MLIT照合 ---",
        f"  差異件数:  {len(comparison)} 件",
        f"  NG件数:    {comp_ng} 件",
        "",
        "--- 期限アラート (120日以内) ---",
        f"  合計:      {len(alerts)} 社",
    ]

    # Alert detail
    for a in alerts:
        lines.append(f"    [{a['stage']:>10}] {a['company_name']} (exp={a['expiry_date']}, {a['days_remaining']}d)")

    lines.extend([
        "",
        "--- DB更新 ---",
        f"  更新件数:  {len(db_updates)} 件",
    ])
    for u in db_updates:
        lines.append(f"    {u['company_name']}: {u['field']} {u['old_value']} -> {u['new_value']} [{u['action']}]")

    lines.extend([
        "",
        "--- 業種カバレッジ (Top 15) ---",
    ])
    for trade, cnt in sorted(all_trades.items(), key=lambda x: -x[1])[:15]:
        lines.append(f"    {trade}: {cnt} 社")

    lines.extend([
        "",
        "--- 出力ファイル ---",
        f"  全社CSV:       {CSV_ALL}",
        f"  照合CSV:       {CSV_COMPARISON}",
        f"  アラートCSV:   {CSV_ALERTS}",
        f"  Excel:         {XLSX_REGISTRY}",
        f"  DB更新ログ:    {CSV_DB_UPDATE_LOG}",
        f"  サマリー:      {TXT_SUMMARY}",
        "=" * 60,
    ])

    summary_text = "\n".join(lines)
    print(summary_text)

    with TXT_SUMMARY.open("w", encoding="utf-8") as f:
        f.write(summary_text)

    print(f"\n  Summary saved: {TXT_SUMMARY}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="MLIT全社スクレイピング + 分析")
    parser.add_argument("--dry-run", action="store_true", help="APIアクセスなし")
    parser.add_argument("--resume", action="store_true", help="中断からの再開")
    args = parser.parse_args()

    start_time = time.time()

    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Content-Type": "application/x-www-form-urlencoded",
    })

    # Resume support
    resume_keys: set[str] = set()
    if args.resume:
        resume_keys = load_resume_keys(CSV_ALL)
        print(f"[RESUME] Loaded {len(resume_keys)} already-processed keys")

    # Phase 1+2: Scrape
    fetch_results = run_scrape(session, resume_keys, dry_run=args.dry_run)

    if args.dry_run:
        print("\n[DRY-RUN] Phases 3-7 skipped")
        return

    if not fetch_results:
        # If resume, reload from CSV
        if args.resume and CSV_ALL.exists():
            print("\n[RESUME] Loading previous results from CSV...")
            fetch_results = []
            with CSV_ALL.open("r", encoding="utf-8-sig") as f:
                for row in csv.DictReader(f):
                    fetch_results.append(FetchRow(
                        company_id=row["company_id"],
                        company_name=row["company_name"],
                        permit_number=row["permit_number"],
                        authority=row["authority"],
                        category=row.get("category", ""),
                        expiry_date=row.get("expiry_date", ""),
                        expiry_wareki=row.get("expiry_wareki", ""),
                        days_remaining=int(row.get("days_remaining", 0)),
                        trades_ippan=row.get("trades_ippan", ""),
                        trades_tokutei=row.get("trades_tokutei", ""),
                        trades_count=int(row.get("trades_count", 0)),
                        fetch_status=row.get("fetch_status", ""),
                        error=row.get("error", ""),
                    ))
            print(f"  Loaded {len(fetch_results)} records from CSV")

    if not fetch_results:
        print("\n[WARN] No results to analyze. Exiting.")
        return

    # Phase 3: Comparison
    comparison = run_comparison(fetch_results)

    # Phase 4: Alerts
    alerts = run_expiry_alerts(fetch_results)

    # Phase 5: Excel
    run_excel(fetch_results, comparison, alerts)

    # Phase 6: DB update
    db_updates = run_db_update(fetch_results)

    # Phase 7: Summary
    run_summary(fetch_results, comparison, alerts, db_updates)

    elapsed = time.time() - start_time
    print(f"\n[DONE] Total elapsed: {elapsed:.0f}s ({elapsed/60:.1f}min)")


if __name__ == "__main__":
    main()
