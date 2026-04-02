"""
export_vendor_ledger.py — 客先テンプレート形式「継続取引業者リスト」Excel出力

145社Excelベースリスト + permit_tracker.db から、客先提出用の
継続取引業者リストを生成する。

Usage:
    python src/export_vendor_ledger.py
"""
from __future__ import annotations

import os
import re
import sqlite3
import sys
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

# ---------------------------------------------------------------------------
# パス定義
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).parent.parent
DB_PATH = PROJECT_ROOT / "data" / "permit_tracker.db"
ORIGINALS_DIR = PROJECT_ROOT / "data" / "originals"
EXCEL_145 = Path("C:/ProgramData/RK10/継続取引業者リスト_145社.xlsx")
OUTPUT_PATH = PROJECT_ROOT / "output" / "継続取引業者リスト_出力.xlsx"

TODAY = date.today()

# ---------------------------------------------------------------------------
# 29業種定義 (略称, 正式名称)
# ---------------------------------------------------------------------------
TRADE_29: list[tuple[str, str]] = [
    ("土", "土木工事業"), ("建", "建築工事業"), ("大", "大工工事業"), ("左", "左官工事業"),
    ("と", "とび・土工工事業"), ("石", "石工事業"), ("屋", "屋根工事業"), ("電", "電気工事業"),
    ("管", "管工事業"), ("タ", "タイル・れんが・ブロック工事業"), ("鋼", "鋼構造物工事業"),
    ("筋", "鉄筋工事業"), ("舗", "舗装工事業"), ("しゅ", "しゅんせつ工事業"),
    ("板", "板金工事業"), ("ガ", "ガラス工事業"), ("塗", "塗装工事業"), ("防", "防水工事業"),
    ("内", "内装仕上工事業"), ("機", "機械器具設置工事業"), ("絶", "熱絶縁工事業"),
    ("通", "電気通信工事業"), ("園", "造園工事業"), ("井", "さく井工事業"),
    ("具", "建具工事業"), ("水", "水道施設工事業"), ("消", "消防施設工事業"),
    ("清", "清掃施設工事業"), ("解", "解体工事業"),
]

# 業種マッチング用キーワード (正式名称から「工事業」を除いた部分で部分一致)
TRADE_MATCH_KEYS: list[str] = [
    name.replace("工事業", "") for _, name in TRADE_29
]

# ---------------------------------------------------------------------------
# 列幅定義
# A-I (9列) + 29業種列(幅3) + AM-AT (8列)
# ---------------------------------------------------------------------------
COL_WIDTHS: list[int] = (
    [30, 10, 12, 12, 28, 4, 4, 8, 12]  # A-I
    + [3] * 29                            # J-AL (29業種)
    + [12, 12, 12, 40, 14, 12, 35, 14]   # AM-AT
)

# ---------------------------------------------------------------------------
# ヘッダ定義
# A-I (9列) + 29業種略称 + AM-AT (8列)
# ---------------------------------------------------------------------------
HEADERS: list[str] = [
    "会社名",               # A
    "代表敬称",             # B
    "代表者名",             # C
    "担当者名",             # D
    "連絡先（メール・電話）",  # E
    "建設業許可番号",        # F (般/特)
    "",                     # G (年)
    "",                     # H (番号)
    "許可区分（知事／大臣）",  # I
] + [abbr for abbr, _ in TRADE_29] + [  # J-AL (29業種)
    "許可満了日",           # AM
    "更新確認日",           # AN
    "許可証受領日",         # AO
    "データ保存場所（フォルダパス）",  # AP
    "現在ステータス",       # AQ
    "書類ステータス",       # AR
    "不足書類",             # AS
    "直近1年発注実績",      # AT
]

# 業種列の開始・終了列インデックス (1-based)
TRADE_COL_START = 10  # J列
TRADE_COL_END = 10 + 29 - 1  # AL列 (38)
# シフト後の列インデックス (1-based)
COL_EXPIRY = TRADE_COL_END + 1      # AM (39)
COL_REVIEW_DATE = TRADE_COL_END + 2  # AN (40)
COL_RECEIPT = TRADE_COL_END + 3      # AO (41)
COL_FOLDER = TRADE_COL_END + 4       # AP (42)
COL_STATUS = TRADE_COL_END + 5       # AQ (43)
COL_DOC_STATUS = TRADE_COL_END + 6   # AR (44)
COL_MISSING_DOCS = TRADE_COL_END + 7 # AS (45)
COL_ORDER = TRADE_COL_END + 8        # AT (46)

# 中央揃え列インデックスセット (1-based) — F,G,H,I + 29業種 + AM,AN,AO,AQ,AR,AT
CENTER_COLS: set[int] = (
    {6, 7, 8, 9}
    | set(range(TRADE_COL_START, TRADE_COL_END + 1))
    | {COL_EXPIRY, COL_REVIEW_DATE, COL_RECEIPT, COL_STATUS, COL_DOC_STATUS, COL_ORDER}
)

# ---------------------------------------------------------------------------
# スタイル定数
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_SIDE = Side(style="thin", color="000000")
THIN_BORDER = Border(
    left=THIN_SIDE, right=THIN_SIDE,
    top=THIN_SIDE, bottom=THIN_SIDE,
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 業種○スタイル
TRADE_CIRCLE_FONT = Font(bold=True, color="4CAF50", size=10)

# ステータス別スタイル
STYLE_UNDER_REVIEW = {"fill": PatternFill("solid", fgColor="FFF2CC")}
STYLE_EXPIRED = {"font": Font(color="FF0000", size=10)}
STYLE_EXPIRING_90 = {"font": Font(color="E65100", size=10)}
STYLE_VALID = {}
STYLE_DOC_INCOMPLETE = {}
STYLE_NOT_SUBMITTED = {"font": Font(color="999999", size=10)}
STYLE_UNIDENTIFIED = {"font": Font(color="999999", size=10)}
STYLE_NO_PERMIT = {"font": Font(color="999999", size=10), "fill": PatternFill("solid", fgColor="F2F2F2")}

# 書類ステータス用スタイル
FONT_DOC_COMPLETE = Font(bold=True, color="4CAF50", size=10)   # 9/9: 緑文字太字
FONT_DOC_INCOMPLETE = Font(color="E65100", size=10)             # 1-8/9: オレンジ文字

FONT_DEFAULT = Font(size=10)


# ---------------------------------------------------------------------------
# 業種マッチング
# ---------------------------------------------------------------------------
def match_trades(trade_names: list[str]) -> list[str]:
    """
    permit_tradesのtrade_name一覧から29業種の○/空欄リストを返す。
    部分一致で判定: TRADE_MATCH_KEYSのキーワードがtrade_nameに含まれれば○。
    """
    result: list[str] = [""] * 29
    for trade_name in trade_names:
        for idx, key in enumerate(TRADE_MATCH_KEYS):
            if key in trade_name:
                result[idx] = "○"
    return result


# ---------------------------------------------------------------------------
# 和暦変換
# ---------------------------------------------------------------------------
def iso_to_wareki(iso_str: str | None) -> str:
    """ISO日付 (YYYY-MM-DD...) → 令和表記 (R7.5.8)。"""
    if not iso_str or iso_str == "UNCERTAIN":
        return ""
    try:
        d = date.fromisoformat(iso_str[:10])
    except (ValueError, TypeError):
        return iso_str
    ry = d.year - 2018
    if ry >= 1:
        return f"R{ry}.{d.month}.{d.day}"
    return iso_str


# ---------------------------------------------------------------------------
# DB接続
# ---------------------------------------------------------------------------
def get_connection() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA foreign_keys = ON")
    conn.row_factory = sqlite3.Row
    return conn


# ---------------------------------------------------------------------------
# DB読み取り
# ---------------------------------------------------------------------------
def load_email_to_company(conn: sqlite3.Connection) -> dict[str, str]:
    """email(小文字) → company_id"""
    rows = conn.execute("SELECT email, company_id FROM company_emails").fetchall()
    result: dict[str, str] = {}
    for r in rows:
        email = r["email"].strip().lower()
        if email not in result:
            result[email] = r["company_id"]
    return result


def load_company_names(conn: sqlite3.Connection) -> dict[str, str]:
    """company_id → official_name"""
    rows = conn.execute("SELECT company_id, official_name FROM companies").fetchall()
    return {r["company_id"]: r["official_name"] for r in rows}


def load_sender_emails(conn: sqlite3.Connection) -> dict[str, str]:
    """company_id → 最初のsenderメアド"""
    rows = conn.execute(
        "SELECT company_id, email FROM company_emails "
        "WHERE email_type = 'sender' ORDER BY email_id"
    ).fetchall()
    result: dict[str, str] = {}
    for r in rows:
        if r["company_id"] not in result:
            result[r["company_id"]] = r["email"]
    return result


def load_permits_by_company(conn: sqlite3.Connection) -> dict[str, list[dict[str, Any]]]:
    """company_id → [permit dict, ...] (current_flag=1のみ)"""
    rows = conn.execute(
        """SELECT p.permit_id, p.company_id, p.permit_number,
                  p.permit_authority, p.permit_category, p.permit_year,
                  p.issue_date, p.expiry_date
           FROM permits p
           WHERE p.current_flag = 1
           ORDER BY p.company_id, p.permit_id"""
    ).fetchall()
    result: dict[str, list[dict[str, Any]]] = {}
    for r in rows:
        d = dict(r)
        result.setdefault(d["company_id"], []).append(d)
    return result


def load_permit_trades(conn: sqlite3.Connection) -> dict[int, list[str]]:
    """permit_id → [trade_name, ...]"""
    rows = conn.execute(
        "SELECT permit_id, trade_name FROM permit_trades ORDER BY trade_id"
    ).fetchall()
    result: dict[int, list[str]] = {}
    for r in rows:
        result.setdefault(r["permit_id"], []).append(r["trade_name"])
    return result


def load_under_review_companies(conn: sqlite3.Connection) -> set[str]:
    """permit_applicationsでstatus='under_review'の会社IDセット"""
    rows = conn.execute(
        "SELECT DISTINCT company_id FROM permit_applications "
        "WHERE status = 'under_review'"
    ).fetchall()
    return {r["company_id"] for r in rows}


def load_receipt_dates(conn: sqlite3.Connection) -> dict[str, str]:
    """company_id → 最初のreceived_at (日付部分)"""
    rows = conn.execute(
        """SELECT company_id, MIN(received_at) as first_received
           FROM receipt_events
           WHERE company_id IS NOT NULL
           GROUP BY company_id"""
    ).fetchall()
    return {r["company_id"]: r["first_received"] for r in rows}


def load_companies_with_receipts(conn: sqlite3.Connection) -> set[str]:
    """receipt_eventsに受信記録がある会社IDセット"""
    rows = conn.execute(
        "SELECT DISTINCT company_id FROM receipt_events "
        "WHERE company_id IS NOT NULL"
    ).fetchall()
    return {r["company_id"] for r in rows}


def load_companies_with_permits_page(conn: sqlite3.Connection) -> set[str]:
    """建設業許可証ページがある会社IDセット"""
    rows = conn.execute(
        "SELECT DISTINCT company_id FROM pages "
        "WHERE company_id IS NOT NULL AND doc_type_name = '建設業許可証'"
    ).fetchall()
    return {r["company_id"] for r in rows}


def load_company_docs(conn: sqlite3.Connection) -> dict[str, set[str]]:
    """company_id → {doc_type_name, ...}"""
    rows = conn.execute(
        "SELECT DISTINCT company_id, doc_type_name FROM pages "
        "WHERE company_id IS NOT NULL AND doc_type_name IS NOT NULL AND doc_type_name != ''"
    ).fetchall()
    result: dict[str, set[str]] = {}
    for r in rows:
        result.setdefault(r["company_id"], set()).add(r["doc_type_name"])
    return result


# ---------------------------------------------------------------------------
# originals/ ディレクトリ名からフォルダパスを取得
# ---------------------------------------------------------------------------
def load_originals_paths() -> dict[str, str]:
    """company_id → 'data/originals/{dirname}/' のパス"""
    result: dict[str, str] = {}
    if not ORIGINALS_DIR.exists():
        return result
    for entry in ORIGINALS_DIR.iterdir():
        if entry.is_dir():
            dirname = entry.name
            # C0008_三和シャッター工業株式会社 刈谷統括営業所 のようなフォーマット
            match = re.match(r"^(C\d+)_", dirname)
            if match:
                cid = match.group(1)
                result[cid] = f"data/originals/{dirname}/"
    return result


# ---------------------------------------------------------------------------
# 145社Excel読み取り
# ---------------------------------------------------------------------------
def load_145_list() -> list[dict[str, Any]]:
    """
    145社Excelを読み込み、辞書リストを返す。
    A列=会社名, E列=メアド を使用。
    """
    wb = openpyxl.load_workbook(str(EXCEL_145), read_only=True)
    ws = wb.active
    result: list[dict[str, Any]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        a_val = str(row[0] or "").strip()
        e_val = str(row[4] or "").strip() if len(row) > 4 else ""
        if not a_val:
            continue

        emails = [e.strip() for e in re.split(r"[\n\r]+", e_val) if e.strip()]

        result.append({
            "company_name": a_val,
            "emails": emails,
        })

    wb.close()
    return result


# ---------------------------------------------------------------------------
# マッチング: 145社 → company_id (重複除去)
# ---------------------------------------------------------------------------
def match_and_deduplicate(
    list_145: list[dict[str, Any]],
    email_to_cid: dict[str, str],
) -> list[dict[str, Any]]:
    """
    145社リストの各行にcompany_idを紐付け、company_id重複は1行に集約。
    company_id=Noneの行はそのまま残す。
    """
    seen_cids: set[str] = set()
    result: list[dict[str, Any]] = []

    for entry in list_145:
        cid = None
        for email in entry["emails"]:
            email_lower = email.lower()
            if email_lower in email_to_cid:
                cid = email_to_cid[email_lower]
                break

        # company_id重複は除外
        if cid and cid in seen_cids:
            continue
        if cid:
            seen_cids.add(cid)

        result.append({
            **entry,
            "company_id": cid,
        })

    return result


# ---------------------------------------------------------------------------
# ステータス判定
# ---------------------------------------------------------------------------
def determine_status(
    company_id: str | None,
    permit: dict[str, Any] | None,
    under_review_cids: set[str],
    companies_with_receipts: set[str],
    has_permit_page: bool,
) -> str:
    """個別permit（または会社全体）のステータスを判定。"""
    if company_id is None:
        return "未特定"

    # 1. 更新申請中
    if company_id in under_review_cids:
        if permit is None:
            return "更新申請中"

    # permit がある場合
    if permit is not None:
        expiry = permit.get("expiry_date")
        if expiry and expiry != "UNCERTAIN":
            try:
                expiry_date = date.fromisoformat(expiry[:10])
                if expiry_date < TODAY:
                    # 申請中の場合は「更新申請中」優先
                    if company_id in under_review_cids:
                        return "更新申請中"
                    return "期限切れ"
                days_remaining = (expiry_date - TODAY).days
                if days_remaining <= 90:
                    return "期限90日以内"
            except (ValueError, TypeError):
                pass
        return "有効"

    # permit がない場合
    if company_id in under_review_cids:
        return "更新申請中"

    # 5. receipt_events受信あり だが permitsなし → 書類不備
    if company_id in companies_with_receipts and not has_permit_page:
        return "書類不備"

    if company_id in companies_with_receipts and has_permit_page:
        # 許可証ページはあるがpermitsレコードがない（OCR未処理等）
        return "書類不備"

    # 6. receipt_events受信なし → 未提出
    return "未提出"


# ---------------------------------------------------------------------------
# Excel出力
# ---------------------------------------------------------------------------
def write_excel(
    matched_list: list[dict[str, Any]],
    company_names: dict[str, str],
    sender_emails: dict[str, str],
    permits_by_company: dict[str, list[dict[str, Any]]],
    permit_trades: dict[int, list[str]],
    under_review_cids: set[str],
    receipt_dates: dict[str, str],
    companies_with_receipts: set[str],
    companies_with_permit_page: set[str],
    originals_paths: dict[str, str],
    company_docs: dict[str, set[str]],
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "継続取引業者リスト"

    # --- Row 1: 凡例行 ---
    # A1: ステータス色分け凡例
    status_legend = (
        "【凡例】 "
        "■有効  "
        "■更新申請中(黄背景)  "
        "■期限切れ(赤文字)  "
        "■期限90日以内(オレンジ文字)  "
        "■書類不備  "
        "■未提出(グレー文字)  "
        "■未特定(グレー文字)  "
        "■許可証なし(グレー背景)  "
        "■9/9提出(緑)  "
        "■不足あり(オレンジ)"
    )
    ws.cell(row=1, column=1, value=status_legend).font = Font(size=9, color="333333", bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    # J1:AL1: 業種略称凡例
    legend_parts = [f"{abbr}={name}" for abbr, name in TRADE_29]
    legend_text = "  ".join(legend_parts)
    legend_cell = ws.cell(row=1, column=TRADE_COL_START, value=legend_text)
    legend_cell.font = Font(size=8, color="333333")
    legend_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws.merge_cells(
        start_row=1, start_column=TRADE_COL_START,
        end_row=1, end_column=TRADE_COL_END,
    )

    # --- Row 2: ヘッダ行 ---
    header_row = 2
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # --- 列幅 ---
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- official_name昇順でソート ---
    def sort_key(entry: dict[str, Any]) -> str:
        cid = entry.get("company_id")
        if cid and cid in company_names:
            return company_names[cid]
        return entry["company_name"]

    sorted_list = sorted(matched_list, key=sort_key)

    # --- データ行 (Row 3~) ---
    row_idx = header_row + 1  # 3
    for entry in sorted_list:
        cid = entry.get("company_id")

        # 表示名
        if cid and cid in company_names:
            display_name = company_names[cid]
        else:
            display_name = entry["company_name"]

        # メアド
        email_display = ""
        if cid and cid in sender_emails:
            email_display = sender_emails[cid]
        elif entry.get("emails"):
            email_display = entry["emails"][0]

        # フォルダパス
        folder_path = originals_paths.get(cid, "") if cid else ""

        # 受領日
        receipt_date_raw = receipt_dates.get(cid, "") if cid else ""
        receipt_date = iso_to_wareki(receipt_date_raw[:10]) if receipt_date_raw else ""

        # 許可情報
        permits = permits_by_company.get(cid, []) if cid else []
        has_permit_page = cid in companies_with_permit_page if cid else False

        # 書類充足情報 (1行目のみ表示)
        ALL_DOCS = [
            "取引申請書", "建設業許可証", "決算書", "会社案内", "工事経歴書",
            "取引先一覧表", "労働安全衛生誓約書", "資格略字一覧", "労働者名簿",
        ]
        docs = company_docs.get(cid, set()) if cid else set()
        has_count = sum(1 for dt in ALL_DOCS if dt in docs)
        doc_status = f"{has_count}/9提出" if cid and has_count > 0 else ""
        missing: list[str] = []
        for dt in ALL_DOCS:
            if dt not in docs:
                missing.append(f"{dt}（任意）" if dt == "会社案内" else dt)
        missing_text = "、".join(missing) if missing and cid and has_count > 0 else ""

        if not permits:
            # 許可情報なしの行（1行出力）
            status = determine_status(
                cid, None, under_review_cids,
                companies_with_receipts, has_permit_page,
            )
            no_permit_label = "許可証なし" if cid else ""
            trade_cols = [""] * 29
            row_data: list[Any] = (
                [display_name, "", "", "", email_display,  # A-E
                 no_permit_label, "", "", ""]               # F-I
                + trade_cols                                # J-AL
                + ["", "", receipt_date, folder_path,       # AM-AP
                   status, doc_status, missing_text, ""]    # AQ-AT
            )
            _write_row(ws, row_idx, row_data, status, no_permit=bool(cid))
            row_idx += 1
        else:
            # 許可情報ありの行（permit毎に1行）
            for i, permit in enumerate(permits):
                trades = permit_trades.get(permit["permit_id"], [])
                trade_cols = match_trades(trades)
                expiry_wareki = iso_to_wareki(permit.get("expiry_date"))

                status = determine_status(
                    cid, permit, under_review_cids,
                    companies_with_receipts, has_permit_page,
                )

                if i == 0:
                    # 1行目: A-E列 + 書類ステータスを表示
                    row_data = (
                        [display_name, "", "", "", email_display,  # A-E
                         permit.get("permit_category", ""),        # F
                         permit.get("permit_year", ""),             # G
                         permit.get("permit_number", ""),           # H
                         permit.get("permit_authority", "")]        # I
                        + trade_cols                                # J-AL
                        + [expiry_wareki, "", receipt_date,         # AM-AO
                           folder_path, status,                     # AP-AQ
                           doc_status, missing_text, ""]            # AR-AT
                    )
                else:
                    # 2行目以降: A-E列・書類ステータスは空
                    row_data = (
                        ["", "", "", "", "",                        # A-E
                         permit.get("permit_category", ""),        # F
                         permit.get("permit_year", ""),             # G
                         permit.get("permit_number", ""),           # H
                         permit.get("permit_authority", "")]        # I
                        + trade_cols                                # J-AL
                        + [expiry_wareki, "", "",                   # AM-AO
                           "", status, "", "", ""]                  # AP-AT
                    )

                _write_row(ws, row_idx, row_data, status)
                row_idx += 1

    # --- freeze + filter ---
    ws.freeze_panes = "A3"  # Row 1=凡例, Row 2=ヘッダ → Row 3以降スクロール
    last_row = row_idx - 1
    last_col_letter = get_column_letter(len(HEADERS))
    if last_row >= header_row + 1:
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_row}"

    # --- 保存 ---
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_PATH))

    return last_row - header_row  # data row count


def _write_row(
    ws: Any,
    row_idx: int,
    row_data: list[Any],
    status: str,
    no_permit: bool = False,
) -> None:
    """1行分のデータをワークシートに書き込み、スタイルを適用。"""
    # ステータス別のスタイルを取得
    style_map = {
        "更新申請中": STYLE_UNDER_REVIEW,
        "期限切れ": STYLE_EXPIRED,
        "期限90日以内": STYLE_EXPIRING_90,
        "有効": STYLE_VALID,
        "書類不備": STYLE_DOC_INCOMPLETE,
        "未提出": STYLE_NOT_SUBMITTED,
        "未特定": STYLE_UNIDENTIFIED,
    }
    if no_permit:
        style = STYLE_NO_PERMIT
    else:
        style = style_map.get(status, {})

    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = THIN_BORDER
        cell.font = FONT_DEFAULT
        cell.alignment = ALIGN_LEFT

        # 中央揃え列
        if col_idx in CENTER_COLS:
            cell.alignment = ALIGN_CENTER

        # 業種列の○スタイル (緑太字)
        if TRADE_COL_START <= col_idx <= TRADE_COL_END and val == "○":
            cell.font = TRADE_CIRCLE_FONT
            # ステータス別fillがある場合はそちらを優先適用
            if "fill" in style:
                cell.fill = style["fill"]
            continue  # fontはTRADE_CIRCLE_FONTを維持

        # 書類ステータス列の色分け (ステータス別fillより優先)
        if col_idx == COL_DOC_STATUS and val:
            if str(val).startswith("9/9"):
                cell.font = FONT_DOC_COMPLETE
            else:
                cell.font = FONT_DOC_INCOMPLETE
            if "fill" in style:
                cell.fill = style["fill"]
            continue

        # ステータス別スタイル適用
        if "fill" in style:
            cell.fill = style["fill"]
        if "font" in style:
            cell.font = style["font"]


# ---------------------------------------------------------------------------
# メイン
# ---------------------------------------------------------------------------
def main() -> None:
    print("=" * 60)
    print(" 継続取引業者リスト Excel出力")
    print("=" * 60)

    # --- DB読み取り ---
    conn = get_connection()
    email_to_cid = load_email_to_company(conn)
    company_names = load_company_names(conn)
    sender_emails = load_sender_emails(conn)
    permits_by_company = load_permits_by_company(conn)
    permit_trades_map = load_permit_trades(conn)
    under_review_cids = load_under_review_companies(conn)
    receipt_dates = load_receipt_dates(conn)
    companies_with_receipts = load_companies_with_receipts(conn)
    companies_with_permit_page = load_companies_with_permits_page(conn)
    company_docs = load_company_docs(conn)
    conn.close()

    # --- originals/ディレクトリ ---
    originals_paths = load_originals_paths()

    print(f"  DB: {len(company_names)}社, {len(sender_emails)}メアド, "
          f"{sum(len(v) for v in permits_by_company.values())}許可証, "
          f"{len(under_review_cids)}申請中")
    print(f"  受信記録: {len(companies_with_receipts)}社, "
          f"フォルダ: {len(originals_paths)}件")

    # --- 145社Excel読み取り ---
    list_145 = load_145_list()
    print(f"  145社Excel: {len(list_145)}行")

    # --- マッチング + 重複除去 ---
    matched = match_and_deduplicate(list_145, email_to_cid)
    matched_count = sum(1 for m in matched if m["company_id"])
    unmatched_count = sum(1 for m in matched if not m["company_id"])
    print(f"  マッチ: {matched_count}社, 未マッチ: {unmatched_count}社")

    # --- Excel出力 ---
    data_rows = write_excel(
        matched,
        company_names,
        sender_emails,
        permits_by_company,
        permit_trades_map,
        under_review_cids,
        receipt_dates,
        companies_with_receipts,
        companies_with_permit_page,
        originals_paths,
        company_docs,
    )

    output_uri = str(OUTPUT_PATH).replace("\\", "/")
    print(f"\n  出力: file:///{output_uri}")
    print(f"  データ行数: {data_rows}")
    print(f"  サイズ: {OUTPUT_PATH.stat().st_size / 1024:.1f} KB")
    print("=" * 60)


if __name__ == "__main__":
    main()
